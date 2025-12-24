import contextlib
import re
from typing import Any, Dict, List, Tuple

import numpy as np
import openpyxl
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.trendline import Trendline
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from domain.constants import BLUE, RED, DataTableColumns, ParamTableColumns
from domain.models import InitialDataItem, InitialDataItemList
from domain.ports import CellDataIO, CellRepository
from domain.utils import (
    calculate_drift,
    calculate_real_custom_area,
    calculate_rn_sqrt,
    calculate_rns,
    calculate_rns_per_sample,
    calculate_square,
    drop_nans,
    linear_fit,
)


def _export_cells_grid(ws_cells, cell_grid_values: List[Tuple[str, str, str]]):
    """Export the 4x4 grid of cell summary to the first sheet (visual aid)."""
    init_data = list(cell_grid_values)
    output = []
    blocks = [init_data[i : i + 4] for i in range(0, len(init_data), 4)]
    for block in blocks:
        block_transposed = list(map(list, zip(*block)))
        output.extend(block_transposed)

    def _maybe_numeric(val):
        if val in (None, ""):
            return None
        if isinstance(val, (int, float)):
            try:
                return int(val) if isinstance(val, int) else round(float(val), 3)
            except Exception:
                return val
        if isinstance(val, str):
            s = val.strip()
            # Values like "Уход: 0.123" or "RnS: 55.3" -> take right part
            if ":" in s:
                s = s.split(":", 1)[1].strip()
            s = s.replace(",", ".")
            # pure number?
            if re.fullmatch(r"-?\d+(?:\.\d+)?", s):
                try:
                    return round(float(s), 3)
                except Exception:
                    return val
        return val

    for row_ind, row in enumerate(output, 1):
        for col_ind, coll in enumerate(row, 1):
            # Header row every 3rd (names) stays text; others (drift, RnS) try numeric
            is_header = (row_ind - 1) % 3 == 0
            value = coll
            if not is_header:
                num = _maybe_numeric(coll)
                value = None if num in (None, "") else num
            cell = ws_cells.cell(row=row_ind, column=col_ind, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # format floats to 3 decimals
            if not is_header and isinstance(value, float):
                cell.number_format = "0.000"
            if (row_ind - 1) % 3 == 0:
                cell.border = Border(
                    right=Side(style="thick"), left=Side(style="thick"), top=Side("thick"), bottom=Side(style="thick")
                )
                cell.font = Font(bold=True)
            elif (row_ind - 3) % 3 == 0:
                cell.border = Border(right=Side(style="thick"), left=Side(style="thick"), bottom=Side(style="thick"))
            else:
                cell.border = Border(right=Side(style="thick"), left=Side(style="thick"))

    for row in ws_cells.rows:
        ws_cells.row_dimensions[row[0].row].height = 21

    # Autofit columns so long cell names/values are fully visible
    def _autofit(col_idx: int, extra: int = 2, min_width: int = 12):
        max_len = 0
        for r in range(1, ws_cells.max_row + 1):
            val = ws_cells.cell(row=r, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        if max_len == 0:
            return
        ws_cells.column_dimensions[get_column_letter(col_idx)].width = max(max_len + extra, min_width)

    for col_idx in range(1, ws_cells.max_column + 1):
        _autofit(col_idx)


_SANITIZE_MAP = {
    "/": "／",  # FULLWIDTH SOLIDUS
    "\\": "＼",  # FULLWIDTH REVERSE SOLIDUS
    ":": "：",  # FULLWIDTH COLON
    "*": "∗",  # ASTERISK OPERATOR
    "?": "？",  # FULLWIDTH QUESTION MARK
    "[": "［",  # FULLWIDTH LEFT SQUARE BRACKET
    "]": "］",  # FULLWIDTH RIGHT SQUARE BRACKET
}
_DESANITIZE_MAP = {v: k for k, v in _SANITIZE_MAP.items()}


def _sanitize_title_component(text: str) -> str:
    if not isinstance(text, str):
        return str(text)
    out = []
    for ch in text:
        out.append(_SANITIZE_MAP.get(ch, ch))
    return "".join(out)


def _desanitize_title_component(text: str) -> str:
    if not isinstance(text, str):
        return str(text)
    out = []
    for ch in text:
        out.append(_DESANITIZE_MAP.get(ch, ch))
    return "".join(out)


def _compose_cell_sheet_title(cell_index: int, name: str, existing: List[str]) -> str:
    prefix = f"Cell №{cell_index} "
    safe_name = _sanitize_title_component(name)
    # Excel sheet title max length = 31
    max_name_len = max(0, 31 - len(prefix))
    base = prefix + safe_name[:max_name_len]
    title = base
    suffix = 2
    while title in existing:
        # try appending _n within 31 chars
        suf = f"_{suffix}"
        title = (base[: max(0, 31 - len(suf))] + suf) if len(base) + len(suf) > 31 else base + suf
        suffix += 1
    return title


class XlsxCellIO(CellDataIO):
    """XLSX adapter implementing combined per-cell sheet with data+results and a chart."""

    def save(self, file_name: str, cell_grid_values: List[Tuple[str, str, str]], repo: CellRepository) -> None:
        wb = openpyxl.Workbook()
        ws_cells = wb.active
        ws_cells.title = "Cells data"
        _export_cells_grid(ws_cells, cell_grid_values)

        # Data columns to export: exclude DRIFT (per request)
        export_data_columns: List[DataTableColumns] = [c for c in DataTableColumns if c is not DataTableColumns.DRIFT]

        def _coerce_value(val, dtype):
            # Convert values to proper numeric types so Excel treats them as numbers
            if val in (None, ""):
                return None
            try:
                if dtype is float:
                    if isinstance(val, str):
                        # Normalize decimal separator: comma -> dot
                        val = val.replace(",", ".")
                    return round(float(val), 3)
                if dtype is int:
                    if isinstance(val, str):
                        val = val.replace(",", ".")
                        # Allow integers provided as "1.0"
                        return int(float(val))
                    return int(val)
                if dtype is bool:
                    if isinstance(val, str):
                        return val in ("True", "TRUE", "true", "1")
                    return bool(val)
            except Exception:
                return val
            return val

        def _is_true(val) -> bool:
            return val in (True, "TRUE", "True", "true", 1, "1")

        def _nonzero_number(val):
            try:
                if val in (None, ""):
                    return None
                num = float(val)
                if num == 0:
                    return None
                return num
            except Exception:
                return None

        # Params to export in results (omit custom diameters)
        results_params: List[ParamTableColumns] = [
            p
            for p in ParamTableColumns
            if p not in (ParamTableColumns.D_CUSTOM1, ParamTableColumns.D_CUSTOM2, ParamTableColumns.D_CUSTOM3)
        ]

        # Create one sheet per recorded cell with data table on the left and results on the right
        for cell_data in repo:
            sheet_name = _compose_cell_sheet_title(cell_data.cell, cell_data.name, wb.sheetnames)
            ws = wb.create_sheet(sheet_name)

            # Write data header (row 1) with styling; widths will be autofit below
            for col_idx, col_def in enumerate(export_data_columns, start=1):
                hcell = ws.cell(row=1, column=col_idx, value=col_def.slug)
                hcell.font = Font(bold=True)
                hcell.border = Border(bottom=Side(style="medium"))
                hcell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data values from InitialDataItemList
            # Determine how many rows are present in initial data
            max_row_index = 0
            for it in cell_data.initial_data:
                if isinstance(it, InitialDataItem):
                    max_row_index = max(max_row_index, it.row)
                else:
                    try:
                        max_row_index = max(max_row_index, it["row"])  # type: ignore[index]
                    except Exception:
                        pass

            # Build mapping from full column enum index to exported column position
            export_col_position = {col.index: pos for pos, col in enumerate(export_data_columns, start=1)}

            # Write rows 2..(max_row_index+2)
            for it in cell_data.initial_data:
                row = it.row if isinstance(it, InitialDataItem) else it["row"]
                col = it.col if isinstance(it, InitialDataItem) else it["col"]
                val = it.value if isinstance(it, InitialDataItem) else it["value"]
                pos = export_col_position.get(col)
                if pos is not None:
                    col_def = next(c for c in export_data_columns if c.index == col)
                    coerced = _coerce_value(val, col_def.dtype)
                    c = ws.cell(row=row + 2, column=pos, value=coerced)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    if col_def.dtype is float and coerced not in (None, ""):
                        c.number_format = "0.000"

            # Results header placed to the right with a gap column
            results_start_col = len(export_data_columns) + 2
            for i, param in enumerate(results_params, start=0):
                hcell = ws.cell(row=1, column=results_start_col + i, value=param.name)
                hcell.font = Font(bold=True)
                hcell.border = Border(bottom=Side(style="medium"))
                hcell.alignment = Alignment(horizontal="center", vertical="center")

                # Value row (2)
                raw_value = getattr(cell_data, param.slug, "")
                value = _coerce_value(raw_value, param.dtype)
                vcell = ws.cell(row=2, column=results_start_col + i, value=value)
                vcell.alignment = Alignment(horizontal="center", vertical="center")
                if param.dtype is float and value not in (None, ""):
                    vcell.number_format = "0.000"

            # Autofit widths for data and results columns based on content
            def _autofit(col_idx: int, extra: int = 2, min_width: int = 10):
                try:
                    max_len = 0
                    for r in range(1, ws.max_row + 1):
                        val = ws.cell(row=r, column=col_idx).value
                        if val is None:
                            continue
                        max_len = max(max_len, len(str(val)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + extra, min_width)
                except Exception:
                    pass

            for col_idx in range(1, len(export_data_columns) + 1):
                _autofit(col_idx)
            for i, _ in enumerate(results_params, start=0):
                _autofit(results_start_col + i)

            # Column mapping for clarity (Excel letters)
            results_row = 2
            data_col_idx = {
                col: export_col_position.get(col.index)
                for col in DataTableColumns
                if export_col_position.get(col.index)
            }
            data_col_letter = {col: get_column_letter(idx) for col, idx in data_col_idx.items()}
            result_col_idx = {param: results_start_col + i for i, param in enumerate(results_params)}
            result_col_letter = {param: get_column_letter(idx) for param, idx in result_col_idx.items()}

            def data_ref(
                col: DataTableColumns, row: int, absolute_col: bool = False, absolute_row: bool = False
            ) -> str:
                letter = data_col_letter.get(col)
                if not letter:
                    return ""
                return f"{'$' if absolute_col else ''}{letter}{'$' if absolute_row else ''}{row}"

            def result_ref(param: ParamTableColumns, row: int = results_row, absolute: bool = True) -> str:
                letter = result_col_letter[param]
                return f"${letter}${row}" if absolute else f"{letter}{row}"

            diameter_col_idx = data_col_idx.get(DataTableColumns.DIAMETER)
            resistance_col_idx = data_col_idx.get(DataTableColumns.RESISTANCE)
            select_col_idx = data_col_idx.get(DataTableColumns.SELECT)
            rn_sqrt_col_idx = data_col_idx.get(DataTableColumns.RN_SQRT)
            rns_col_idx = data_col_idx.get(DataTableColumns.RNS)
            rns_error_col_idx = data_col_idx.get(DataTableColumns.RNS_ERROR)
            square_col_idx = data_col_idx.get(DataTableColumns.SQUARE)

            def _row_has_selected_data(row: int) -> bool:
                if diameter_col_idx is None or resistance_col_idx is None:
                    return False
                if select_col_idx:
                    sel_val = ws.cell(row=row, column=select_col_idx).value
                    if not _is_true(sel_val):
                        return False
                diam_val = _nonzero_number(ws.cell(row=row, column=diameter_col_idx).value)
                res_val = _nonzero_number(ws.cell(row=row, column=resistance_col_idx).value)
                return diam_val is not None and res_val is not None

            data_max_row = 1
            for r in range(2, max_row_index + 2):
                if _row_has_selected_data(r):
                    data_max_row = max(data_max_row, r)
            if data_max_row < 2:
                data_max_row = 2

            slope_ref = result_ref(ParamTableColumns.SLOPE)
            intercept_ref = result_ref(ParamTableColumns.INTERCEPT)
            drift_ref = result_ref(ParamTableColumns.DRIFT)
            rns_res_ref = result_ref(ParamTableColumns.RNS)
            rn_const_ref = result_ref(ParamTableColumns.RN_CONSISTENT)
            s_custom1_ref = result_ref(ParamTableColumns.S_CUSTOM1)
            s_custom2_ref = result_ref(ParamTableColumns.S_CUSTOM2)
            s_custom3_ref = result_ref(ParamTableColumns.S_CUSTOM3)
            planned_drift_ref = result_ref(ParamTableColumns.PLANNED_DRIFT)

            rn_sqrt_range = (
                f"{data_col_letter[DataTableColumns.RN_SQRT]}2:{data_col_letter[DataTableColumns.RN_SQRT]}{data_max_row}"
                if DataTableColumns.RN_SQRT in data_col_letter
                else None
            )
            diameter_range = (
                f"{data_col_letter[DataTableColumns.DIAMETER]}2:{data_col_letter[DataTableColumns.DIAMETER]}{data_max_row}"
                if DataTableColumns.DIAMETER in data_col_letter
                else None
            )

            # Results formulas (slope/intercept/drift/RnS/errors/real areas) with IFERROR guards
            if rn_sqrt_range and diameter_range:
                slope_cell = ws.cell(row=results_row, column=result_col_idx[ParamTableColumns.SLOPE])
                slope_cell.value = (
                    f'=IF(COUNT({rn_sqrt_range})<2,"",IFERROR(SLOPE({rn_sqrt_range},{diameter_range}),""))'
                )
                slope_cell.number_format = "0.0000"

                intercept_cell = ws.cell(row=results_row, column=result_col_idx[ParamTableColumns.INTERCEPT])
                intercept_cell.value = (
                    f'=IF(COUNT({rn_sqrt_range})<2,"",IFERROR(INTERCEPT({rn_sqrt_range},{diameter_range}),""))'
                )
                intercept_cell.number_format = "0.0000"

            drift_cell = ws.cell(row=results_row, column=result_col_idx[ParamTableColumns.DRIFT])
            drift_cell.value = (
                f'=IF(OR(ISBLANK({slope_ref}),{slope_ref}=0,ISBLANK({intercept_ref})),"",'
                f'IFERROR(-{intercept_ref}/{slope_ref},""))'
            )
            drift_cell.number_format = "0.000"

            rns_cell = ws.cell(row=results_row, column=result_col_idx[ParamTableColumns.RNS])
            rns_cell.value = f'=IF(OR(ISBLANK({slope_ref}),{slope_ref}=0),"",IFERROR(PI()*0.25/({slope_ref}^2),""))'
            rns_cell.number_format = "0.000"

            if rns_col_idx:
                rns_range = (
                    f"{data_col_letter[DataTableColumns.RNS]}2:{data_col_letter[DataTableColumns.RNS]}{data_max_row}"
                )
                select_range = None
                if select_col_idx:
                    select_range = f"{data_col_letter[DataTableColumns.SELECT]}2:{data_col_letter[DataTableColumns.SELECT]}{data_max_row}"
                rns_error_cell = ws.cell(row=results_row, column=result_col_idx[ParamTableColumns.RNS_ERROR])
                if select_range:
                    cond = f"(({select_range}=TRUE)*ISNUMBER({rns_range})*({rns_range}<>0))"
                else:
                    cond = f"(ISNUMBER({rns_range})*({rns_range}<>0))"
                cnt = f"SUMPRODUCT({cond})"
                mean = f"SUMPRODUCT({cond}*{rns_range})/{cnt}"
                variance_sum = f"SUMPRODUCT({cond}*({rns_range}-{mean})^2)"
                std_expr = f"SQRT({variance_sum}/{cnt})"
                rns_error_cell.value = f'=IF({cnt}=0,"",IFERROR({std_expr},""))'
                rns_error_cell.number_format = "0.000"

            s_real_c1_cell = ws.cell(row=results_row, column=result_col_idx[ParamTableColumns.S_REAL_CUSTOM1])
            s_real_c1_cell.value = (
                f'=IF(ISBLANK({s_custom1_ref}),"",'
                f'IFERROR(PI()/4*(SQRT(4*{s_custom1_ref}/PI())+{planned_drift_ref}-{drift_ref})^2,""))'
            )
            s_real_c1_cell.number_format = "0.000"

            s_real_c2_cell = ws.cell(row=results_row, column=result_col_idx[ParamTableColumns.S_REAL_CUSTOM2])
            s_real_c2_cell.value = (
                f'=IF(ISBLANK({s_custom2_ref}),"",'
                f'IFERROR(PI()/4*(SQRT(4*{s_custom2_ref}/PI())+{planned_drift_ref}-{drift_ref})^2,""))'
            )
            s_real_c2_cell.number_format = "0.000"

            s_real_c3_cell = ws.cell(row=results_row, column=result_col_idx[ParamTableColumns.S_REAL_CUSTOM3])
            s_real_c3_cell.value = (
                f'=IF(ISBLANK({s_custom3_ref}),"",'
                f'IFERROR(PI()/4*(SQRT(4*{s_custom3_ref}/PI())+{planned_drift_ref}-{drift_ref})^2,""))'
            )
            s_real_c3_cell.number_format = "0.000"

            # Per-row formulas for derived values (Rn^-0.5, RnS, площадь, ошибка RnS)
            for r in range(2, data_max_row + 1):
                if not _row_has_selected_data(r):
                    continue
                ws.cell(row=r, column=diameter_col_idx).value if diameter_col_idx else None
                ws.cell(row=r, column=resistance_col_idx).value if resistance_col_idx else None
                select_ref = None
                if select_col_idx:
                    select_ref = f"${get_column_letter(select_col_idx)}{r}"

                if rn_sqrt_col_idx and resistance_col_idx:
                    res_ref = data_ref(DataTableColumns.RESISTANCE, r)
                    base = f"1/SQRT({res_ref}+{rn_const_ref})"
                    core = f'IFERROR({base},"")'
                    if select_ref:
                        formula = f'=IF({select_ref},{core},"")'
                    else:
                        formula = f'=IF(OR(ISBLANK({res_ref}),ISBLANK({rn_const_ref})),"",{core})'
                    cell = ws.cell(row=r, column=rn_sqrt_col_idx)
                    cell.value = formula
                    cell.number_format = "0.000"

                if square_col_idx and diameter_col_idx:
                    diam_ref = data_ref(DataTableColumns.DIAMETER, r)
                    base = f"({diam_ref}-{drift_ref})^2*PI()/4"
                    core = f'IFERROR({base},"")'
                    if select_ref:
                        formula = f'=IF({select_ref},{core},"")'
                    else:
                        formula = f'=IF(OR(ISBLANK({diam_ref}),ISBLANK({drift_ref})),"",{core})'
                    cell = ws.cell(row=r, column=square_col_idx)
                    cell.value = formula
                    cell.number_format = "0.000"

                if rns_col_idx and resistance_col_idx and diameter_col_idx:
                    diam_ref = data_ref(DataTableColumns.DIAMETER, r)
                    res_ref = data_ref(DataTableColumns.RESISTANCE, r)
                    base = f"({res_ref}+{rn_const_ref})*PI()/4*({diam_ref}-{drift_ref})^2"
                    core = f'IFERROR({base},"")'
                    if select_ref:
                        formula = f'=IF({select_ref},{core},"")'
                    else:
                        formula = f'=IF(OR(ISBLANK({res_ref}),ISBLANK({diam_ref}),ISBLANK({drift_ref})),"",{core})'
                    cell = ws.cell(row=r, column=rns_col_idx)
                    cell.value = formula
                    cell.number_format = "0.000"

                if rns_error_col_idx and rns_col_idx:
                    rns_ref_row = data_ref(DataTableColumns.RNS, r)
                    base = f"ABS({rns_ref_row}-{rns_res_ref})"
                    core = f'IFERROR({base},"")'
                    if select_ref:
                        formula = f'=IF({select_ref},{core},"")'
                    else:
                        formula = f'=IF(ISBLANK({rns_ref_row}),"",{core})'
                    cell = ws.cell(row=r, column=rns_error_col_idx)
                    cell.value = formula
                    cell.number_format = "0.000"

            # Build chart: Rn^-0.5 vs Диаметр ACAD (μm)
            try:
                # Find exported columns for RN_SQRT and DIAMETER
                y_col_idx = next(i for i, c in enumerate(export_data_columns, start=1) if c is DataTableColumns.RN_SQRT)
                x_col_idx = next(
                    i for i, c in enumerate(export_data_columns, start=1) if c is DataTableColumns.DIAMETER
                )

                # Determine data max row based on last non-empty in x/y columns
                def last_non_empty_row_for_col(col_idx: int) -> int:
                    last = 1
                    for r in range(2, data_max_row + 1):
                        v = ws.cell(row=r, column=col_idx).value
                        if v not in (None, ""):
                            last = r
                    return last

                max_row = last_non_empty_row_for_col(y_col_idx)
                if max_row < 2:
                    max_row = 2

                chart = ScatterChart()
                chart.title = "Rn^-0.5 vs Диаметр ACAD (μm)"
                # Show only markers (points) to avoid spurious lines/series
                chart.scatterStyle = "marker"
                chart.style = 2
                chart.x_axis.title = "Диаметр ACAD (μm)"
                chart.y_axis.title = "Rn^-0.5"
                # Title/legend and axes styling
                with contextlib.suppress(Exception):
                    chart.legend.position = "r"  # right, outside plot area
                    chart.legend.overlay = False
                    # put chart title outside overlay
                    if hasattr(chart, "title") and hasattr(chart.title, "overlay"):
                        chart.title.overlay = False
                    # Axis positions and numeric display
                    chart.x_axis.axPos = "b"
                    chart.y_axis.axPos = "l"
                    chart.x_axis.delete = False
                    chart.y_axis.delete = False
                    chart.x_axis.number_format = "General"
                    chart.y_axis.number_format = "General"
                    chart.x_axis.majorTickMark = "out"
                    chart.y_axis.majorTickMark = "out"
                    # Ensure tick labels appear next to axes and not skipped
                    chart.x_axis.tickLblPos = "nextTo"
                    chart.y_axis.tickLblPos = "nextTo"
                    with contextlib.suppress(Exception):
                        chart.x_axis.tickLblSkip = 1
                        chart.x_axis.tickMarkSkip = 1
                        chart.y_axis.tickLblSkip = 1
                        chart.y_axis.tickMarkSkip = 1
                    if getattr(chart.x_axis, "title", None):
                        chart.x_axis.title.overlay = False
                    if getattr(chart.y_axis, "title", None):
                        chart.y_axis.title.overlay = False
                # Gray grid and show tick labels next to axes
                with contextlib.suppress(Exception):
                    grid_color = "B3B3B3"  # light gray
                    for axis in (chart.x_axis, chart.y_axis):
                        axis.tickLblPos = "nextTo"
                        axis.majorGridlines = ChartLines()
                        axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=grid_color))

                xvalues = Reference(ws, min_col=x_col_idx, min_row=2, max_row=max_row)
                yvalues = Reference(ws, min_col=y_col_idx, min_row=2, max_row=max_row)
                series = Series(yvalues, xvalues, title="Data")
                # Uniform marker style & color; hide connecting line
                series.marker = Marker(symbol="circle")
                series.marker.size = 3
                with contextlib.suppress(Exception):
                    blue = BLUE.lstrip("#").upper()
                    series.marker.graphicalProperties.solidFill = blue
                    series.marker.graphicalProperties.line.solidFill = blue
                    if getattr(series.graphicalProperties, "line", None) is not None:
                        series.graphicalProperties.line.noFill = True
                # Trendline: show equation only (no R^2); paint it RED
                with contextlib.suppress(Exception):
                    series.trendline = Trendline(trendlineType="linear", dispEq=True, dispRSqr=False)
                    red = RED.lstrip("#").upper()
                    try:
                        series.trendline.graphicalProperties.line.solidFill = red
                    except Exception:
                        # Fallback property name in some versions
                        series.trendline.spPr = GraphicalProperties(ln=LineProperties(solidFill=red))

                chart.varyColors = False
                chart.series.append(series)

                # Ensure axis range includes x-intercept (drift); extend trendline via forecast to reach it
                with contextlib.suppress(Exception):
                    drift = float(getattr(cell_data, "drift", 0.0))
                    x_vals: List[float] = []
                    for r in range(2, max_row + 1):
                        xv = ws.cell(row=r, column=x_col_idx).value
                        if xv not in (None, ""):
                            x_vals.append(float(xv))
                    if x_vals:
                        chart.x_axis.scaling.min = min(min(x_vals), drift)
                        chart.x_axis.scaling.max = max(max(x_vals), drift)
                        # Forecast distances measured along X units
                        try:
                            backward = max(0.0, min(x_vals) - drift)
                            forward = max(0.0, drift - max(x_vals))
                            if hasattr(series.trendline, "backward"):
                                series.trendline.backward = round(backward, 2)
                            if hasattr(series.trendline, "forward"):
                                series.trendline.forward = round(forward, 2)
                        except Exception:
                            pass

                # Place chart under the results table and stretch to Q21
                chart_anchor_row = 4
                chart_anchor_col = results_start_col
                start_marker = AnchorMarker(col=chart_anchor_col - 1, colOff=0, row=chart_anchor_row - 1, rowOff=0)
                end_marker = AnchorMarker(col=18 - 1, row=23 - 1, colOff=0, rowOff=0)  # bottom-right at Q21
                chart.anchor = TwoCellAnchor(_from=start_marker, to=end_marker, editAs="twoCell")
                ws.add_chart(chart)
            except Exception:
                # Chart creation should not break export
                pass

        if not file_name.endswith(".xlsx"):
            file_name += ".xlsx"
        wb.save(filename=file_name)

    def load(self, file_name: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        wb = openpyxl.load_workbook(file_name)

        # Combined-sheet format: sheets starting with "Cell №"
        combined_sheet_names = [sh for sh in wb.sheetnames if sh.startswith("Cell №")]
        if combined_sheet_names:
            return self._load_combined(wb, combined_sheet_names)

        # Fallback to legacy format (separate Data/Results sheets)
        from .persistence_xlsx import load_cells_from_xlsx as _legacy_load  # lazy import to avoid cycles

        return _legacy_load(file_name)

    def _load_combined(self, wb, sheet_names: List[str]) -> Tuple[List[Dict[str, Any]], List[str]]:
        items: List[Dict[str, Any]] = []
        errors: List[str] = []

        for sheet_name in sheet_names:
            try:
                i = int(re.findall(r"Cell №(\d+) ", sheet_name)[0])
                cell_name = re.findall(r"Cell №\d+ (.*)", sheet_name)[0]
                # de-sanitize sheet component back to original visible ASCII if needed
                cell_name = _desanitize_title_component(cell_name)
            except (IndexError, ValueError):
                errors.append(f"Неверное имя листа: {sheet_name}")
                continue

            ws = wb[sheet_name]

            # Map header titles to all column indexes (1-based)
            header_positions: Dict[str, List[int]] = {}
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=c).value
                if isinstance(val, str) and val.strip():
                    key = val.strip()
                    header_positions.setdefault(key, []).append(c)

            def first_col_for(name: str) -> int | None:
                pos = header_positions.get(name)
                return pos[0] if pos else None

            def last_col_for(name: str) -> int | None:
                pos = header_positions.get(name)
                return pos[-1] if pos else None

            header_aliases: Dict[ParamTableColumns, List[str]] = {
                ParamTableColumns.S_REAL_CUSTOM1: ["S_1.00 (μm²)"],
            }

            def last_col_for_param(param: ParamTableColumns) -> int | None:
                names = [param.name] + header_aliases.get(param, [])
                for nm in names:
                    col = last_col_for(nm)
                    if col:
                        return col
                return None

            def read_cell(row: int, col: int | None) -> Any:
                return ws.cell(row=row, column=col).value if col is not None else None

            def to_float(val, default=None):
                if val in (None, "", "None"):
                    return default
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    s = val.replace(",", ".").strip()
                    try:
                        return float(s)
                    except Exception:
                        return default
                return default

            def to_bool(val) -> bool:
                if val in (True, "TRUE", "True", "true", 1, "1"):
                    return True
                return False

            def last_non_empty_row_for_col(col_idx: int | None) -> int:
                if col_idx is None:
                    return 1
                last = 1
                for r in range(2, ws.max_row + 1):
                    v = ws.cell(row=r, column=col_idx).value
                    if v not in (None, ""):
                        last = r
                return last

            diameter_col = first_col_for(DataTableColumns.DIAMETER.slug)
            resistance_col = first_col_for(DataTableColumns.RESISTANCE.slug)
            select_col = first_col_for(DataTableColumns.SELECT.slug)
            name_col = first_col_for(DataTableColumns.NAME.slug)
            number_col = first_col_for(DataTableColumns.NUMBER.slug)

            if diameter_col is None or resistance_col is None:
                errors.append(f"Колонки с исходными данными не найдены в листе {sheet_name}")
                continue

            candidate_cols = [c for c in (diameter_col, resistance_col, name_col, number_col) if c]
            max_data_row = 2
            for c in candidate_cols:
                max_data_row = max(max_data_row, last_non_empty_row_for_col(c))

            # Input parameters stored in the results table (row 2)
            def read_param(param: ParamTableColumns, default=None):
                col = last_col_for_param(param)
                return to_float(read_cell(2, col), default)

            rn_consistent = read_param(ParamTableColumns.RN_CONSISTENT, default=0.0)
            allowed_error = read_param(ParamTableColumns.ALLOWED_ERROR, default=2.5)
            s_custom1 = read_param(ParamTableColumns.S_CUSTOM1, default=1.0)
            s_custom2 = read_param(ParamTableColumns.S_CUSTOM2, default=1.0)
            s_custom3 = read_param(ParamTableColumns.S_CUSTOM3, default=1.0)
            planned_drift = read_param(ParamTableColumns.PLANNED_DRIFT, default=1.0)

            rows_data: List[Dict[str, Any]] = []
            for row in range(2, max_data_row + 1):
                row_idx = row - 2
                selected_raw = read_cell(row, select_col)
                row_dict = {
                    "row": row_idx,
                    "number": read_cell(row, number_col) or row_idx + 1,
                    "name": read_cell(row, name_col) or "",
                    "selected": to_bool(selected_raw),
                    "diameter": to_float(read_cell(row, diameter_col)),
                    "resistance": to_float(read_cell(row, resistance_col)),
                }
                if not row_dict["selected"] and any(
                    v is not None for v in (row_dict["diameter"], row_dict["resistance"])
                ):
                    # Treat rows with numeric data as selected even if checkbox is empty
                    row_dict["selected"] = True
                rows_data.append(row_dict)

            # Calculate derived values based on parsed raw data
            diameter_for_calc: List[float] = []
            rn_sqrt_for_calc: List[float] = []
            for rd in rows_data:
                if not rd.get("selected"):
                    continue
                if rd.get("diameter") is None or rd.get("resistance") is None:
                    continue
                with contextlib.suppress(Exception):
                    rn_sqrt_val = calculate_rn_sqrt(
                        resistance=float(rd["resistance"]), rn_consistent=rn_consistent or 0.0
                    )
                    rd["rn_sqrt"] = rn_sqrt_val
                    diameter_for_calc.append(float(rd["diameter"]))
                    rn_sqrt_for_calc.append(float(rn_sqrt_val))

            slope = intercept = drift = rns = 0.0
            try:
                diam_arr, rn_arr = drop_nans(diameter_for_calc, rn_sqrt_for_calc)
            except Exception:
                diam_arr, rn_arr = np.array([], dtype=float), np.array([], dtype=float)

            def _is_nan(val: Any) -> bool:
                try:
                    return bool(np.isnan(float(val)))
                except Exception:
                    return False

            if len(diam_arr) >= 2:
                with contextlib.suppress(Exception):
                    slope, intercept = linear_fit(diam_arr, rn_arr)
            if _is_nan(slope):
                slope = 0.0
            if _is_nan(intercept):
                intercept = 0.0
            if slope not in (None, "", 0):
                with contextlib.suppress(Exception):
                    drift = calculate_drift(slope=slope, intercept=intercept)
                with contextlib.suppress(Exception):
                    rns = calculate_rns(slope)

            rns_values: List[float] = []
            for rd in rows_data:
                if not rd.get("selected"):
                    continue
                if rd.get("diameter") is None or rd.get("resistance") is None:
                    continue
                with contextlib.suppress(Exception):
                    rd["square"] = calculate_square(diameter=float(rd["diameter"]), drift=float(drift))
                with contextlib.suppress(Exception):
                    rns_val = calculate_rns_per_sample(
                        resistance=float(rd["resistance"]),
                        diameter=float(rd["diameter"]),
                        drift=float(drift),
                        rn_persistent=rn_consistent or 0.0,
                    )
                    rd["rns"] = rns_val
                    rns_values.append(float(rns_val))

            if rns_values:
                try:
                    rns_error = float(np.sqrt(np.sum((np.array(rns_values, dtype=float) - rns) ** 2) / len(rns_values)))
                except Exception:
                    rns_error = 0.0
            else:
                rns_error = 0.0

            for rd in rows_data:
                if not rd.get("selected"):
                    continue
                if "rns" in rd:
                    with contextlib.suppress(Exception):
                        rd["rns_error"] = abs(float(rd["rns"]) - float(rns))

            def calc_area(area_nominal):
                with contextlib.suppress(Exception):
                    return calculate_real_custom_area(
                        area_nominal=area_nominal,
                        planned_drift=float(planned_drift or 0.0),
                        drift=float(drift),
                    )
                return 0.0

            s_real_c1 = calc_area(s_custom1)
            s_real_c2 = calc_area(s_custom2)
            s_real_c3 = calc_area(s_custom3)

            initial_data = InitialDataItemList()

            def add_item(row_idx: int, col: DataTableColumns, value: Any):
                v = "" if value in (None, "") else value
                initial_data.append(InitialDataItem(row=row_idx, col=col.index, value=v))

            for rd in rows_data:
                row_idx = rd["row"]
                add_item(row_idx, DataTableColumns.NUMBER, rd.get("number") or row_idx + 1)
                add_item(row_idx, DataTableColumns.NAME, rd.get("name") or "")
                add_item(row_idx, DataTableColumns.SELECT, "True" if rd.get("selected") else "")
                add_item(row_idx, DataTableColumns.DIAMETER, rd.get("diameter"))
                add_item(row_idx, DataTableColumns.RESISTANCE, rd.get("resistance"))
                add_item(row_idx, DataTableColumns.RNS, rd.get("rns"))
                add_item(row_idx, DataTableColumns.RNS_ERROR, rd.get("rns_error"))
                add_item(row_idx, DataTableColumns.DRIFT, drift if rd.get("selected") else "")
                add_item(row_idx, DataTableColumns.SQUARE, rd.get("square"))
                add_item(row_idx, DataTableColumns.RN_SQRT, rd.get("rn_sqrt"))

            items.append(
                {
                    "cell": i,
                    "name": cell_name,
                    "diameter_list": [
                        rd.get("diameter") for rd in rows_data if rd.get("selected") and rd.get("diameter") is not None
                    ],
                    "rn_sqrt_list": [
                        rd.get("rn_sqrt") for rd in rows_data if rd.get("selected") and rd.get("rn_sqrt") is not None
                    ],
                    "slope": slope,
                    "intercept": intercept,
                    "drift": drift,
                    "rns": rns,
                    "drift_error": 0.0,
                    "rns_error": rns_error,
                    "initial_data": initial_data,
                    "rn_consistent": rn_consistent if rn_consistent is not None else 0.0,
                    "allowed_error": allowed_error if allowed_error is not None else 0.0,
                    "s_custom1": s_custom1 if s_custom1 is not None else 0.0,
                    "s_custom2": s_custom2 if s_custom2 is not None else 0.0,
                    "s_custom3": s_custom3 if s_custom3 is not None else 0.0,
                    "planned_drift": planned_drift if planned_drift is not None else 0.0,
                    "s_real_custom1": s_real_c1,
                    "s_real_custom2": s_real_c2,
                    "s_real_custom3": s_real_c3,
                }
            )

        return items, errors
