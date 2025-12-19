from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font

from domain.constants import DataTableColumns, ParamTableColumns
from domain.models import InitialDataItem, InitialDataItemList

_SANITIZE_MAP = {
    "/": "／",
    "\\": "＼",
    ":": "：",
    "*": "∗",
    "?": "？",
    "[": "［",
    "]": "］",
}


def _sanitize_sheet_name(name: str) -> str:
    safe = "".join(_SANITIZE_MAP.get(ch, ch) for ch in str(name))
    return safe[:31] if len(safe) > 31 else safe


def _to_float(val):
    if val in (None, "", "None"):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val.replace(",", "."))
        except Exception:
            return None
    return None


def save_template(file_path: str, sheet_name: str, rows: List[Dict], areas: Dict[str, float | None]) -> str:
    """Create template workbook with headers and optional area/diameter columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(sheet_name)

    headers = [
        DataTableColumns.NUMBER.slug,
        DataTableColumns.NAME.slug,
        DataTableColumns.SELECT.slug,
        DataTableColumns.DIAMETER.slug,
    ]

    param_columns: List[Tuple[ParamTableColumns, float | None]] = [
        (ParamTableColumns.S_CUSTOM1, areas.get("s_custom1")),
        (ParamTableColumns.S_CUSTOM2, areas.get("s_custom2")),
        (ParamTableColumns.S_CUSTOM3, areas.get("s_custom3")),
        (ParamTableColumns.D_CUSTOM1, areas.get("d_custom1")),
        (ParamTableColumns.D_CUSTOM2, areas.get("d_custom2")),
        (ParamTableColumns.D_CUSTOM3, areas.get("d_custom3")),
    ]
    # Only keep params that were provided (non-None)
    param_columns = [(p, v) for p, v in param_columns if v is not None]

    for col_idx, name in enumerate(headers + [p.name for p, _ in param_columns], start=1):
        hcell = ws.cell(row=1, column=col_idx, value=name)
        hcell.font = Font(bold=True)
        hcell.alignment = Alignment(horizontal="center", vertical="center")

    # Put parameter values in row 2 under their headers
    if param_columns:
        for offset, (_, value) in enumerate(param_columns, start=len(headers) + 1):
            cell = ws.cell(row=2, column=offset, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.000"

    start_row = 3
    for i, row in enumerate(rows):
        ws.cell(row=start_row + i, column=1, value=row.get("number"))
        ws.cell(row=start_row + i, column=2, value=row.get("name"))
        ws.cell(row=start_row + i, column=3, value=bool(row.get("selected", False)))
        dcell = ws.cell(row=start_row + i, column=4, value=row.get("diameter"))
        dcell.number_format = "0.000"
        dcell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=start_row + i, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=start_row + i, column=3).alignment = Alignment(horizontal="center", vertical="center")

    if not file_path.endswith(".xlsx"):
        file_path += ".xlsx"
    wb.save(file_path)
    return file_path


def load_template(file_path: str) -> Tuple[InitialDataItemList, Dict[str, float | None], List[str]]:
    """Read template and return initial_data plus area/diameter values."""
    errors: List[str] = []
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header_positions: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if isinstance(val, str) and val.strip():
            header_positions[val.strip()] = c

    def col_for(slug: str) -> int | None:
        return header_positions.get(slug)

    number_col = col_for(DataTableColumns.NUMBER.slug)
    name_col = col_for(DataTableColumns.NAME.slug)
    select_col = col_for(DataTableColumns.SELECT.slug)
    diameter_col = col_for(DataTableColumns.DIAMETER.slug)

    if diameter_col is None or name_col is None or number_col is None or select_col is None:
        errors.append("Не найдены все обязательные колонки (№, Имя, ✓, Диаметр).")

    def last_non_empty(col_idx: int | None) -> int:
        if col_idx is None:
            return 2
        last = 2
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=col_idx).value not in (None, ""):
                last = r
        return last

    max_row = max(
        last_non_empty(number_col),
        last_non_empty(name_col),
        last_non_empty(select_col),
        last_non_empty(diameter_col),
    )

    initial_data = InitialDataItemList()
    for r in range(3, max_row + 1):
        row_idx = r - 3
        num_val = ws.cell(row=r, column=number_col).value if number_col else ""
        name_val = ws.cell(row=r, column=name_col).value if name_col else ""
        select_val = ws.cell(row=r, column=select_col).value if select_col else ""
        diam_val = ws.cell(row=r, column=diameter_col).value if diameter_col else ""

        if all(v in (None, "", False) for v in (name_val, select_val, diam_val)):
            continue

        initial_data.append(InitialDataItem(row=row_idx, col=DataTableColumns.NUMBER.index, value=num_val))
        initial_data.append(InitialDataItem(row=row_idx, col=DataTableColumns.NAME.index, value=name_val or ""))
        initial_data.append(
            InitialDataItem(
                row=row_idx,
                col=DataTableColumns.SELECT.index,
                value="True" if select_val in (True, "TRUE", "True", "true", 1, "1") else "",
            )
        )
        initial_data.append(InitialDataItem(row=row_idx, col=DataTableColumns.DIAMETER.index, value=diam_val or ""))

    param_values: Dict[str, float | None] = {
        "s_custom1": None,
        "s_custom2": None,
        "s_custom3": None,
        "d_custom1": None,
        "d_custom2": None,
        "d_custom3": None,
    }
    for param, key in (
        (ParamTableColumns.S_CUSTOM1, "s_custom1"),
        (ParamTableColumns.S_CUSTOM2, "s_custom2"),
        (ParamTableColumns.S_CUSTOM3, "s_custom3"),
        (ParamTableColumns.D_CUSTOM1, "d_custom1"),
        (ParamTableColumns.D_CUSTOM2, "d_custom2"),
        (ParamTableColumns.D_CUSTOM3, "d_custom3"),
    ):
        col = col_for(param.name)
        if col:
            param_values[key] = _to_float(ws.cell(row=2, column=col).value)

    return initial_data, param_values, errors
