import re
from typing import List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from domain.constants import DataTableColumns, ParamTableColumns
from domain.models import InitialDataItem, InitialDataItemList
from domain.ports import CellRepository


def save_cells_to_xlsx(
    file_name: str,
    cell_grid_values: List[Tuple[str, str, str]],
    repo: CellRepository,
    data_headers,
    results_headers,
):
    def _autofit(ws, extra: int = 2, min_width: int = 12):
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for r in range(1, ws.max_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            if max_len == 0:
                continue
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + extra, min_width)

    wb = openpyxl.Workbook()
    ws_cells = wb.active
    ws_cells.title = "Cells data"

    init_data = list(cell_grid_values)
    output = []

    blocks = [init_data[i : i + 4] for i in range(0, len(init_data), 4)]
    for block in blocks:
        block_transposed = list(map(list, zip(*block)))
        output.extend(block_transposed)

    for row_ind, row in enumerate(output, 1):
        for col_ind, coll in enumerate(row, 1):
            cell = ws_cells.cell(row=row_ind, column=col_ind, value=coll)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if (row_ind - 1) % 3 == 0:
                cell.border = Border(
                    right=Side(style="thick"), left=Side(style="thick"), top=Side("thick"), bottom=Side(style="thick")
                )
                cell.font = Font(bold=True)
            elif (row_ind - 3) % 3 == 0:
                cell.border = Border(right=Side(style="thick"), left=Side(style="thick"), bottom=Side(style="thick"))
            else:
                cell.border = Border(right=Side(style="thick"), left=Side(style="thick"))

    for row in ws_cells.rows:
        ws_cells.row_dimensions[row[0].row].height = 21
    _autofit(ws_cells)

    for cell_data in repo:
        ws_data = wb.create_sheet(f"Data №{cell_data.cell} {cell_data.name}")
        ws_data.append(data_headers)
        for dat in cell_data.initial_data:
            ws_data.cell(row=dat["row"] + 2, column=dat["col"] + 1, value=dat["value"])
        _autofit(ws_data)

        ws_results = wb.create_sheet(f"Results №{cell_data.cell} {cell_data.name}")
        ws_results.append(results_headers)
        for i, param in enumerate(ParamTableColumns):
            ws_results.cell(row=2, column=i + 1, value=getattr(cell_data, param.slug, ""))
        _autofit(ws_results)

    if not file_name.endswith(".xlsx"):
        file_name += ".xlsx"
    wb.save(filename=file_name)


def load_cells_from_xlsx(file_name: str):
    wb = openpyxl.load_workbook(file_name)
    data_sheet_names = [sh for sh in wb.sheetnames if sh.startswith("Data №")]
    if not len(data_sheet_names):
        raise ValueError("Не найдены данные с нумерацией для записанных ячеек")
    sheet_names = "\n".join(wb.sheetnames)

    items = []
    errors = []

    for sheet_name in data_sheet_names:
        try:
            i = int(re.findall(r"Data №(\d+) .*", sheet_name)[0])
            data_name = re.findall(f"(Data №{i} .*)", sheet_name)[0]
            result_name = re.findall(f"(Results №{i} .*)", sheet_names)[0]
            cell_name = re.findall(f"Data №{i} (.*)", sheet_name)[0]
        except (IndexError, ValueError):
            errors.append(f"Неверное имя листа: {sheet_name}")
            continue

        ws_data = wb[data_name]
        ws_result = wb[result_name]

        initial_data = InitialDataItemList()
        column_names = [ws_data[1][col].value for col in range(0, ws_data.max_column)]
        for data_column in DataTableColumns:
            col = None
            try:
                col = column_names.index(data_column.slug)
            except (ValueError,):
                errors.append(f"Колонка '{data_column.slug}' не найдена в таблице '{data_name}'")

            for row in range(2, ws_data.max_row + 1):
                try:
                    value = ws_data[row][col].value if col is not None and ws_data[row][col].value else ""
                    initial_data.append(InitialDataItem(row=row - 2, col=data_column.index, value=value))
                except IndexError:
                    errors.append(f"Ошибка чтения ячейки row={row} col={col} в '{data_name}'")
                    continue

        diameter_list = [
            float(v.value) if v.value else None for v in initial_data.filter(col=DataTableColumns.DIAMETER.index)
        ]
        rn_sqrt_list = [
            float(v.value) if v.value else None for v in initial_data.filter(col=DataTableColumns.RN_SQRT.index)
        ]

        result_column_names = [ws_result[1][col].value for col in range(0, ws_result.max_column)]
        result_kwargs = {}
        header_aliases = {
            ParamTableColumns.S_REAL_CUSTOM1: ["S_1.00 (μm²)"],
            ParamTableColumns.RN_CONSISTENT: ["Последовательное Rn"],
        }
        optional_slugs = {
            ParamTableColumns.S_REAL_CUSTOM1.slug,
            ParamTableColumns.S_REAL_CUSTOM2.slug,
            ParamTableColumns.S_REAL_CUSTOM3.slug,
            ParamTableColumns.S_CUSTOM1.slug,
            ParamTableColumns.S_CUSTOM2.slug,
            ParamTableColumns.S_CUSTOM3.slug,
            ParamTableColumns.D_CUSTOM1.slug,
            ParamTableColumns.D_CUSTOM2.slug,
            ParamTableColumns.D_CUSTOM3.slug,
        }

        def find_column(param: ParamTableColumns) -> int | None:
            names = [param.name] + header_aliases.get(param, [])
            for nm in names:
                try:
                    return result_column_names.index(nm)
                except (ValueError,):
                    continue
            return None

        for result_column in ParamTableColumns:
            col = find_column(result_column)
            if col is None:
                # Optional new columns may be absent in older files
                if result_column.slug in optional_slugs:
                    continue
                errors.append(f"Колонка '{result_column.name}' не найдена в таблице '{result_name}'")
                continue
            row = 2
            try:
                value = ws_result[row][col].value if col is not None and ws_result[row][col].value else 0
                result_kwargs[result_column.slug] = value
            except IndexError:
                errors.append(f"Ошибка чтения '{result_column.name}' в '{result_name}'")
                continue

        items.append(
            dict(
                cell=i,
                name=cell_name,
                diameter_list=diameter_list,
                rn_sqrt_list=rn_sqrt_list,
                initial_data=initial_data,
                **result_kwargs,
            )
        )

    return items, errors
