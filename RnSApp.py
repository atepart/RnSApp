import sys
import numpy as np
import openpyxl
from PyQt5 import QtWidgets, QtCore, QtGui
import pyqtgraph as pg
from openpyxl.styles import Side, Border, Font

from utils import (
    linear,
    linear_fit,
    calculate_rns,
    calculate_rns_per_sample,
    calculate_drift,
)


class DataTableColumns:
    NUMBER = "№"  # int
    NAME = "Имя"  # str
    DRIFT = "Уход"  # float
    RNS = "RnS"  # float
    DIAMETER = "Диаметр ACAD (μm)"  # float
    RESISTANCE = "Сопротивление (Ω)"  # float
    RN = "Rn^-0.5"  # float

    fields = [
        NUMBER,
        NAME,
        DRIFT,
        RNS,
        DIAMETER,
        RESISTANCE,
        RN,
    ]

    text_by_index = {ind: field for ind, field in enumerate(fields)}
    index_by_text = {field: ind for ind, field in enumerate(fields)}


class ParamTableColumns:
    SLOPE = "Наклон"
    INTERCEPT = "Пересечение"
    DRIFT = "Уход"
    RNS = "RnS"
    DRIFT_ERROR = "Ошибка ухода"
    RNS_ERROR = "Ошибка RnS"

    fields = [
        SLOPE,
        INTERCEPT,
        DRIFT,
        RNS,
        DRIFT_ERROR,
        RNS_ERROR,
    ]

    text_by_index = {ind: field for ind, field in enumerate(fields)}
    index_by_text = {field: ind for ind, field in enumerate(fields)}


class DataTable(QtWidgets.QTableWidget):
    def __init__(self, rows):
        super(DataTable, self).__init__(rows, len(DataTableColumns.fields))

        # Set Table headers
        self.setHorizontalHeaderLabels(DataTableColumns.fields)
        self.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.ResizeMode.Stretch
        )

        # Remove vertical Table headers
        self.verticalHeader().setVisible(False)

        # Set Table grid
        self.setShowGrid(True)
        self.setGridStyle(QtCore.Qt.PenStyle.SolidLine)

        # Set default Numbers
        self.set_default_numbers()

        # Set columns RnS, Rn as read-only
        self.set_read_only_columns(
            [
                DataTableColumns.index_by_text[DataTableColumns.RNS],
                DataTableColumns.index_by_text[DataTableColumns.RN],
            ]
        )

        # Connect event update_table
        self.itemChanged.connect(self.update_table)

    def set_default_numbers(self):
        for i in range(self.rowCount()):
            self.setItem(
                i,
                DataTableColumns.index_by_text[DataTableColumns.NUMBER],
                QtWidgets.QTableWidgetItem(str(i + 1)),
            )

    def set_read_only_columns(self, columns):
        for col in columns:
            self.setItemDelegateForColumn(col, ReadOnlyDelegate(self))

    def update_table(self, item):
        self.itemChanged.disconnect(self.update_table)
        row = item.row()
        col = item.column()

        # Для рассчета нужны только колонки Diameter, Resistance
        if col not in (
            DataTableColumns.index_by_text[DataTableColumns.DIAMETER],
            DataTableColumns.index_by_text[DataTableColumns.RESISTANCE],
        ):
            self.itemChanged.connect(self.update_table)
            return

        # Достаем Resistance
        if (
            self.item(row, DataTableColumns.index_by_text[DataTableColumns.RESISTANCE])
            is None
        ):
            self.itemChanged.connect(self.update_table)
            return

        # Убеждаемся, что Resistance is float
        try:
            resistance = float(
                self.item(
                    row, DataTableColumns.index_by_text[DataTableColumns.RESISTANCE]
                ).text()
            )
        except ValueError:
            self.itemChanged.connect(self.update_table)
            return

        if resistance != 0:  # Если Resistance != 0, то рассчитываем Rn
            rn_sqrt = 1 / np.sqrt(resistance)
            self.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.RN],
                QtWidgets.QTableWidgetItem(str(round(rn_sqrt, 4))),
            )
        else:  # Иначе очищаем RnS, Rn
            self.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.RNS],
                QtWidgets.QTableWidgetItem(""),
            )  # Clear RnS column
            self.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.RN],
                QtWidgets.QTableWidgetItem(""),
            )  # Clear Rn^-0.5 column
        self.itemChanged.connect(self.update_table)

    def keyPressEvent(self, event):
        # На нажатие Enter/Return переход на следующую строку
        if event.key() in (QtCore.Qt.Key.Key_Enter, QtCore.Qt.Key.Key_Return):
            row = self.currentRow()
            col = self.currentColumn()
            if row < self.rowCount() - 1:
                self.setCurrentCell(row + 1, col)

        # На нажатие Delete/Backspace удаление выбранных значений
        elif event.key() in (QtCore.Qt.Key.Key_Delete, QtCore.Qt.Key.Key_Backspace):
            selected_items = self.selectedItems()
            if selected_items:
                for item in selected_items:
                    if item.column() not in [
                        DataTableColumns.index_by_text[DataTableColumns.RNS],
                        DataTableColumns.index_by_text[DataTableColumns.RN],
                    ]:  # Disable delete for columns Rn and RnS
                        self.setItem(
                            item.row(), item.column(), QtWidgets.QTableWidgetItem("")
                        )
                self.parent().update_plot()

        # Ивент вставки ctrl-v
        elif event.matches(QtGui.QKeySequence.Paste):
            self.paste_data()
        else:
            super(DataTable, self).keyPressEvent(event)

    def paste_data(self):
        clipboard = QtWidgets.QApplication.clipboard()
        data = clipboard.text()
        rows = data.split("\n")
        start_row = self.currentRow()
        start_col = self.currentColumn()
        if start_col not in [
            DataTableColumns.index_by_text[DataTableColumns.DIAMETER],
            DataTableColumns.index_by_text[DataTableColumns.RESISTANCE],
            DataTableColumns.index_by_text[DataTableColumns.NUMBER],
            DataTableColumns.index_by_text[DataTableColumns.NAME],
        ]:  # Можно вставлять только в Number, Name, Diameter, Resistance

            return
        for i, row in enumerate(rows):
            values = row.split("\t")
            for j, value in enumerate(values):
                if start_col in [
                    DataTableColumns.index_by_text[DataTableColumns.DIAMETER],
                    DataTableColumns.index_by_text[DataTableColumns.RESISTANCE],
                ]:  # Для данных колонок нужны числа float
                    value = value.replace(",", ".")
                item = QtWidgets.QTableWidgetItem(value)
                self.setItem(start_row + i, start_col + j, item)
                self.update_table(item)

    def get_column_values(self, column: str):
        column_ind = DataTableColumns.index_by_text[column]
        values = []
        for row in range(self.rowCount()):
            value = self.item(row, column_ind)
            if not value and not value.text():
                continue
            values.append(value.text())


class ParamTable(QtWidgets.QTableWidget):
    def __init__(self):
        super(ParamTable, self).__init__(1, len(ParamTableColumns.fields))

        self.setHorizontalHeaderLabels(ParamTableColumns.fields)
        self.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)

        self.verticalHeader().setVisible(False)

        self.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred
        )


class ReadOnlyDelegate(QtWidgets.QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        return


class Window(QtWidgets.QWidget):
    def __init__(self):
        super(Window, self).__init__()
        # Таблица с исходными данными
        self.table = DataTable(rows=50)

        # График
        self.plot = pg.PlotWidget()
        self.prepare_plot()

        # Main Layout
        self.layout = QtWidgets.QHBoxLayout()

        # Right Layout (таблица с параметрами, график, экшнс)
        self.right_layout = QtWidgets.QVBoxLayout()

        # Таблица с параметрами
        self.param_table = ParamTable()

        # Экшнс кнопки
        self.actions_group = QtWidgets.QGroupBox("Действия")
        self.actions_layout = QtWidgets.QHBoxLayout()

        self.result_button = QtWidgets.QPushButton("Result")
        self.result_button.setToolTip("Произвести рассчет")
        self.result_button.clicked.connect(self.calculate_results)

        self.clean_rn_button = QtWidgets.QPushButton("Clear Rn")
        self.clean_rn_button.setToolTip("Очистить Rn")
        self.clean_rn_button.clicked.connect(self.clean_rn)

        self.clean_all_button = QtWidgets.QPushButton("Clear All")
        self.clean_all_button.setToolTip("Очистить все данные")
        self.clean_all_button.clicked.connect(self.clean_all)

        self.save_button = QtWidgets.QPushButton("Save All data")
        self.save_button.setToolTip("Сохранить входные данные и рассчет")
        self.save_button = QtWidgets.QPushButton("Save")
        self.save_button.clicked.connect(self.save_data)

        self.actions_layout.addWidget(self.result_button)
        self.actions_layout.addWidget(self.clean_rn_button)
        self.actions_layout.addWidget(self.clean_all_button)
        self.actions_layout.addWidget(self.save_button)
        self.actions_group.setLayout(self.actions_layout)

        # Грид с кнопками
        self.cell_group = QtWidgets.QGroupBox("Запись")
        self.cell_layout = QtWidgets.QGridLayout()
        self.cell_buttons = []
        for i in range(4):
            for j in range(4):
                button = QtWidgets.QPushButton(f"Ячейка {i * 4 + j + 1}")
                button.clicked.connect(
                    lambda checked=False, row=i, col=j: self.cell_button_clicked(
                        row, col
                    )
                )
                self.cell_layout.addWidget(button, i, j)
                self.cell_buttons.append(button)
        self.cell_save_button = QtWidgets.QPushButton("Save cells RnS")
        self.cell_save_button.setToolTip("Сохранить выходную таблицу с RnS")
        self.cell_save_button.clicked.connect(self.save_cell_data)
        self.cell_layout.addWidget(self.cell_save_button, 4, 3)
        self.cell_group.setLayout(self.cell_layout)

        # Добавляем все виджеты в правый лейаут
        self.right_layout.addWidget(self.param_table)
        self.right_layout.addWidget(self.plot)
        self.right_layout.addWidget(self.cell_group)
        self.right_layout.addWidget(self.actions_group)

        # Добавляет виджеты в основной лейаут
        self.layout.addWidget(self.table)
        self.layout.addLayout(self.right_layout)
        self.setLayout(self.layout)

    def calculate_results(self):
        # Update the RnS column in the table
        self.update_plot()
        for row in range(self.table.rowCount()):
            item_diameter = self.table.item(
                row, DataTableColumns.index_by_text[DataTableColumns.DIAMETER]
            )
            item_rn = self.table.item(
                row, DataTableColumns.index_by_text[DataTableColumns.RESISTANCE]
            )
            if (
                item_diameter is not None
                and item_diameter.text()
                and item_rn is not None
                and item_rn.text()
            ):
                try:
                    diameter = float(item_diameter.text())
                    rn = float(item_rn.text())
                    zero_x = float(
                        self.param_table.item(
                            0, ParamTableColumns.index_by_text[ParamTableColumns.DRIFT]
                        ).text()
                    )  # Get the Уход value from the parameter table
                    rns_value = calculate_rns_per_sample(
                        resistance=rn, diameter=diameter, zero_x=zero_x
                    )
                    self.table.setItem(
                        row,
                        DataTableColumns.index_by_text[DataTableColumns.RNS],
                        QtWidgets.QTableWidgetItem(str(round(rns_value, 4))),
                    )
                    drift_value = calculate_drift(
                        diameter=diameter, resistance=rn, rns=rns_value
                    )
                    self.table.setItem(
                        row,
                        DataTableColumns.index_by_text[DataTableColumns.DRIFT],
                        QtWidgets.QTableWidgetItem(str(round(drift_value, 4))),
                    )
                except ValueError:
                    self.table.setItem(
                        row,
                        DataTableColumns.index_by_text[DataTableColumns.RNS],
                        QtWidgets.QTableWidgetItem(""),
                    )
                    self.table.setItem(
                        row,
                        DataTableColumns.index_by_text[DataTableColumns.DRIFT],
                        QtWidgets.QTableWidgetItem(""),
                    )
                    self.table.setItem(
                        row,
                        DataTableColumns.index_by_text[DataTableColumns.RN],
                        QtWidgets.QTableWidgetItem(""),
                    )
            else:
                self.table.setItem(
                    row,
                    DataTableColumns.index_by_text[DataTableColumns.RNS],
                    QtWidgets.QTableWidgetItem(""),
                )
                self.table.setItem(
                    row,
                    DataTableColumns.index_by_text[DataTableColumns.DRIFT],
                    QtWidgets.QTableWidgetItem(""),
                )
                self.table.setItem(
                    row,
                    DataTableColumns.index_by_text[DataTableColumns.RN],
                    QtWidgets.QTableWidgetItem(""),
                )

        self.update_plot()

    def update_plot(self):
        diameter_list = []
        rn_list = []
        rns_list = []
        drift_list = []
        for i in range(self.table.rowCount()):
            item_diameter = self.table.item(
                i, DataTableColumns.index_by_text[DataTableColumns.DIAMETER]
            )
            item_rn = self.table.item(
                i, DataTableColumns.index_by_text[DataTableColumns.RN]
            )
            item_rns = self.table.item(
                i, DataTableColumns.index_by_text[DataTableColumns.RNS]
            )
            item_drift = self.table.item(
                i, DataTableColumns.index_by_text[DataTableColumns.DRIFT]
            )
            if (
                item_diameter is not None
                and item_rn is not None
                and item_diameter.text()
                and item_rn.text()
            ):
                try:
                    diameter_list.append(float(item_diameter.text()))
                    rn_list.append(float(item_rn.text()))
                except ValueError:
                    pass
        self.plot.clear()
        self.plot.plot(
            diameter_list, rn_list, name="Data", symbolSize=6, symbolBrush="#088F8F"
        )
        if len(diameter_list) > 1 and len(rn_list) > 1:
            slope, intercept = linear_fit(diameter_list, rn_list)

            self.param_table.setItem(
                0,
                ParamTableColumns.index_by_text[ParamTableColumns.SLOPE],
                QtWidgets.QTableWidgetItem(str(round(slope, 4))),
            )
            self.param_table.setItem(
                0,
                ParamTableColumns.index_by_text[ParamTableColumns.INTERCEPT],
                QtWidgets.QTableWidgetItem(str(round(intercept, 4))),
            )

            drift = -intercept / slope

            self.param_table.setItem(
                0,
                ParamTableColumns.index_by_text[ParamTableColumns.DRIFT],
                QtWidgets.QTableWidgetItem(str(round(drift, 4))),
            )

            rns = calculate_rns(slope)
            self.param_table.setItem(
                0,
                ParamTableColumns.index_by_text[ParamTableColumns.RNS],
                QtWidgets.QTableWidgetItem(str(round(rns, 4))),
            )

            rns_list = np.array(
                [
                    float(
                        self.table.item(
                            row, DataTableColumns.index_by_text[DataTableColumns.RNS]
                        ).text()
                    )
                    for row in range(self.table.rowCount())
                    if self.table.item(
                        row, DataTableColumns.index_by_text[DataTableColumns.RNS]
                    )
                    is not None
                ]
            )
            rns_error = np.sqrt(np.sum((rns_list - rns) ** 2))

            self.param_table.setItem(
                0,
                ParamTableColumns.index_by_text[ParamTableColumns.RNS_ERROR],
                QtWidgets.QTableWidgetItem(str(round(rns_error, 4))),
            )

            drift_list = np.array(
                [
                    float(
                        self.table.item(
                            row, DataTableColumns.index_by_text[DataTableColumns.DRIFT]
                        ).text()
                    )
                    for row in range(self.table.rowCount())
                    if self.table.item(
                        row, DataTableColumns.index_by_text[DataTableColumns.DRIFT]
                    )
                    is not None
                ]
            )
            drift_error = np.sqrt(np.sum((drift_list - drift) ** 2))
            self.param_table.setItem(
                0,
                ParamTableColumns.index_by_text[ParamTableColumns.DRIFT_ERROR],
                QtWidgets.QTableWidgetItem(str(round(drift_error, 4))),
            )

            # Plot fit
            if np.min(diameter_list) > drift:
                diameter_list.insert(0, drift)
            if np.max(diameter_list) < drift:
                diameter_list.append(drift)
            y_appr = np.vectorize(lambda x: linear(x, slope, intercept))(diameter_list)
            pen2 = pg.mkPen(color="#FF0000", width=3)
            self.plot.plot(
                diameter_list,
                y_appr,
                name="Fit",
                pen=pen2,
                symbolSize=0,
                symbolBrush=pen2.color(),
            )

    def prepare_plot(self):
        y_label = "Rn^-0.5"
        x_label = "Diameter"
        plot_title = "Rn^-0.5(Diameter)"
        self.plot.setBackground("w")
        self.plot.setTitle(plot_title, color="#413C58", size="10pt")
        styles = {"color": "#413C58", "font-size": "15px"}
        self.plot.setLabel("left", y_label, **styles)
        self.plot.setLabel("bottom", x_label, **styles)
        self.plot.addLegend()
        self.plot.showGrid(x=True, y=True)

    def save_data(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        file_name, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Data",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options,
        )
        if not file_name:
            return
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        headers = [
            self.table.horizontalHeaderItem(i).text()
            for i in range(self.table.columnCount())
        ]
        ws1.append(headers)
        for row in range(self.table.rowCount()):
            data = [
                self.table.item(row, col).text() if self.table.item(row, col) else ""
                for col in range(self.table.columnCount())
            ]
            ws1.append(data)
        ws2 = wb.create_sheet("Results")
        headers = [
            self.param_table.horizontalHeaderItem(i).text()
            for i in range(self.param_table.columnCount())
        ]
        ws2.append(headers)
        for row in range(self.param_table.rowCount()):
            data = [
                self.param_table.item(row, col).text()
                if self.param_table.item(row, col)
                else ""
                for col in range(self.param_table.columnCount())
            ]
            ws2.append(data)

        if not file_name.endswith(".xlsx"):
            file_name += ".xlsx"
        wb.save(filename=file_name)

    def clean_rn(self):
        for row in range(self.table.rowCount()):
            self.table.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.RNS],
                QtWidgets.QTableWidgetItem(""),
            )  # Clear RnS column
            self.table.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.RESISTANCE],
                QtWidgets.QTableWidgetItem(""),
            )  # Clear Resistance column
            self.table.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.RN],
                QtWidgets.QTableWidgetItem(""),
            )  # Clear Rn^-0.5 column
        self.plot.clear()

    def clean_all(self):
        for row in range(self.table.rowCount()):
            self.table.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.NAME],
                QtWidgets.QTableWidgetItem(""),
            )  # Clear Name column
            self.table.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.RNS],
                QtWidgets.QTableWidgetItem(""),
            )  # Clear RnS column
            self.table.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.DIAMETER],
                QtWidgets.QTableWidgetItem(""),
            )  # Clear Diameter column
            self.table.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.RESISTANCE],
                QtWidgets.QTableWidgetItem(""),
            )  # Clear Resistance column
            self.table.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.RN],
                QtWidgets.QTableWidgetItem(""),
            )  # Clear Rn^-0.5 column
            self.table.setItem(
                row,
                DataTableColumns.index_by_text[DataTableColumns.NUMBER],
                QtWidgets.QTableWidgetItem(str(row + 1)),
            )
        self.plot.clear()

    def cell_button_clicked(self, row, col):
        series, ok = QtWidgets.QInputDialog.getText(
            self, "Input Dialog", "Enter series:"
        )
        if ok:
            drift = (
                self.param_table.item(
                    0, ParamTableColumns.index_by_text[ParamTableColumns.DRIFT]
                ).text()
                if self.param_table.item(
                    0, ParamTableColumns.index_by_text[ParamTableColumns.DRIFT]
                )
                is not None
                else ""
            )
            rns = (
                self.param_table.item(
                    0, ParamTableColumns.index_by_text[ParamTableColumns.RNS]
                ).text()
                if self.param_table.item(
                    0, ParamTableColumns.index_by_text[ParamTableColumns.RNS]
                )
                is not None
                else ""
            )
            button = self.cell_buttons[row * 4 + col]
            button.setText(f"Серия: {series}\nУход: {drift}\nRnS: {rns}")

    def save_cell_data(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        file_name, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Cell Data",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options,
        )
        if not file_name:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cell Data"

        def parse_btn_text(btn):
            text = btn.text()
            series = text.split("\n")[0].split(": ")[1] if "Серия" in text else text
            care = text.split("\n")[1].split(": ")[1] if "Уход" in text else ""
            rns = text.split("\n")[2].split(": ")[1] if "RnS" in text else ""
            return [series, care, rns]

        init_data = [parse_btn_text(btn) for btn in self.cell_buttons]
        output = []

        # Разделение исходного массива на блоки по 4 строки
        blocks = [init_data[i : i + 4] for i in range(0, len(init_data), 4)]

        # Обработка каждого блока
        for block in blocks:
            # Перестановка элементов в блоке
            block_transposed = list(map(list, zip(*block)))
            # Добавление переставленного блока в выходной массив
            output.extend(block_transposed)
        for row_ind, row in enumerate(output, 1):
            for col_ind, coll in enumerate(row, 1):
                ws.cell(row=row_ind, column=col_ind, value=coll)
                if (row_ind - 1) % 3 == 0:  # для клеток с названием серии
                    ws.cell(row=row_ind, column=col_ind).border = Border(
                        right=Side(style="thick"),
                        top=Side("thick"),
                        bottom=Side(style="thick"),
                    )
                    ws.cell(row=row_ind, column=col_ind).font = Font(bold=True)
                else:  # для остальных клеток
                    ws.cell(row=row_ind, column=col_ind).border = Border(
                        right=Side(style="thick")
                    )
        # Save the Excel file
        if not file_name.endswith(".xlsx"):
            file_name += ".xlsx"
        wb.save(filename=file_name)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = Window()
    window.show()
    sys.exit(app.exec_())
