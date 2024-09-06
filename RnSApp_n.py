import sys
import numpy as np
import math
import openpyxl
from PyQt5 import QtWidgets, QtCore, QtGui
import pyqtgraph as pg
from utils import linear, linear_fit

class BackgroundDelegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, color, parent=None):
        super(BackgroundDelegate, self).__init__(parent)
        self.color = color

    def paint(self, painter, option, index):
        painter.save()
        painter.fillRect(option.rect, self.color)
        painter.restore()
        super(BackgroundDelegate, self).paint(painter, option, index)

class Table(QtWidgets.QTableWidget):
    def __init__(self, rows, columns):
        super(Table, self).__init__(rows, columns)
        self.setHorizontalHeaderLabels(['Number', 'Name', 'RnS', 'Diameter (μm)', 'Resistance (Ω)', 'Rn^-0.5'])
        self.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.verticalHeader().setVisible(False)
        self.setItemDelegateForColumn(0, ReadOnlyDelegate(self))
        self.itemChanged.connect(self.update_table)
        for i in range(rows):
            self.setItem(i, 0, QtWidgets.QTableWidgetItem(str(i+1)))
        self.setShowGrid(True)
        self.setGridStyle(QtCore.Qt.SolidLine)

        # Set columns 0, 2, 4, and 5 as read-only
        self.set_read_only_columns([0, 2, 5])

        # Set background color for Diameter and Resistance columns
        diameter_column = 3
        resistance_column = 4
        diameter_delegate = self.itemDelegateForColumn(diameter_column)
        resistance_delegate = self.itemDelegateForColumn(resistance_column)
        if diameter_delegate is not None:
            diameter_delegate.setBackground(QtGui.QBrush(QtGui.QColor(134, 255, 170, 128)))
        if resistance_delegate is not None:
            resistance_delegate.setBackground(QtGui.QBrush(QtGui.QColor(134, 255, 170, 128)))

    def set_read_only_columns(self, columns):
        for col in columns:
            if col != 4:  # Make the Resistance column editable
                delegate = ReadOnlyDelegate(self)
                self.setItemDelegateForColumn(col, delegate)

    def update_table(self, item):
        self.itemChanged.disconnect(self.update_table)
        row = item.row()
        col = item.column()
        try:
            if col == 3:  # Diameter
                if self.item(row, 4) is not None:
                    resistance = float(self.item(row, 4).text())
                    if resistance != 0:  # Check if resistance is not equal to 0
                        rn_sqrt = 1 / np.sqrt(resistance)
                        self.setItem(row, 5, QtWidgets.QTableWidgetItem(str(round(rn_sqrt, 4))))
                    else:
                        self.setItem(row, 2, QtWidgets.QTableWidgetItem(''))  # Clear RnS column
                        self.setItem(row, 5, QtWidgets.QTableWidgetItem(''))  # Clear Rn^-0.5 column
            elif col == 4:  # Resistance
                resistance = float(item.text())
                if resistance != 0:  # Check if resistance is not equal to 0
                    rn_sqrt = 1 / np.sqrt(resistance)
                    self.setItem(row, 5, QtWidgets.QTableWidgetItem(str(round(rn_sqrt, 4))))
                else:
                    self.setItem(row, 2, QtWidgets.QTableWidgetItem(''))  # Clear RnS column
                    self.setItem(row, 5, QtWidgets.QTableWidgetItem(''))  # Clear Rn^-0.5 column
        except ValueError:
            pass
        self.itemChanged.connect(self.update_table)

    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_Enter or event.key() == QtCore.Qt.Key_Return:
            row = self.currentRow()
            col = self.currentColumn()
            if row < self.rowCount() - 1:
                self.setCurrentCell(row + 1, col)
        elif event.key() == QtCore.Qt.Key_Delete:
            selected_items = self.selectedItems()
            if selected_items:
                for item in selected_items:
                    if item.column() not in [0, 2, 5]:  # Disable delete for columns 0, 2, and 5
                        self.setItem(item.row(), item.column(), QtWidgets.QTableWidgetItem(''))
                self.parent().update_plot()
        elif event.matches(QtGui.QKeySequence.Paste):
            self.paste_data()
        else:
            super(Table, self).keyPressEvent(event)

    def paste_data(self):
        clipboard = QtWidgets.QApplication.clipboard()
        data = clipboard.text()
        rows = data.split('\n')
        start_row = self.currentRow()
        start_col = self.currentColumn()
        for i, row in enumerate(rows):
            values = row.split('\t')
            for j, value in enumerate(values):
                if start_col in [2, 3, 4, 5]:  # Для данных колонок нужны числа float
                    value = value.replace(',', '.')
                item = QtWidgets.QTableWidgetItem(value)
                self.setItem(start_row + i, start_col + j, item)
                self.update_table(item)
        self.parent().update_plot()

class ReadOnlyDelegate(QtWidgets.QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        return

class Window(QtWidgets.QWidget):
    def __init__(self):
        super(Window, self).__init__()
        self.table = Table(50, 6)
        self.plot = pg.PlotWidget()
        self.y_label = "Rn^-0.5"
        self.x_label = "Diameter"
        self.plot_title = "Rn^-0.5(Diameter)"
        self.prepare_plot()
        self.layout = QtWidgets.QHBoxLayout()
        self.right_layout = QtWidgets.QVBoxLayout()
        self.param_table = QtWidgets.QTableWidget(5, 2)
        self.param_table.setHorizontalHeaderLabels(['Parameter', 'Value'])
        self.param_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.param_table.verticalHeader().setVisible(False)
        self.param_table.setItemDelegateForColumn(0, ReadOnlyDelegate(self))
        self.right_layout.addWidget(self.param_table)
        self.right_layout.addWidget(self.plot)
        self.button_layout = QtWidgets.QHBoxLayout()
        self.result_button = QtWidgets.QPushButton('Result')
        self.result_button.clicked.connect(self.calculate_results)
        self.result_button.setStyleSheet("background-color: rgba(72, 250, 215, 51); border-radius: 10px; border: none;")
        self.clean_rn_button = QtWidgets.QPushButton('Clear Rn')
        self.clean_rn_button.clicked.connect(self.clean_rn)
        self.clean_rn_button.setStyleSheet("background-color: rgba(72, 250, 215, 51); border-radius: 10px; border: none;")
        self.clean_all_button = QtWidgets.QPushButton('Clear All')
        self.clean_all_button.clicked.connect(self.clean_all)
        self.clean_all_button.setStyleSheet("background-color: rgba(72, 250, 215, 51); border-radius: 10px; border: none;")
        self.save_button = QtWidgets.QPushButton('Save')
        self.save_button.clicked.connect(self.save_data)
        self.save_button.setStyleSheet("background-color: rgba(72, 250, 215, 51); border-radius: 10px; border: none;")
        self.button_layout.addWidget(self.result_button)
        self.button_layout.addWidget(self.clean_rn_button)
        self.button_layout.addWidget(self.clean_all_button)
        self.button_layout.addWidget(self.save_button)
        self.right_layout.addLayout(self.button_layout)
        self.result_table = QtWidgets.QTableWidget(1, 6)
        self.result_table.setHorizontalHeaderLabels(['Number', 'Name', 'RnS', 'Diameter (μm)', 'Resistance (Ω)', 'Rn^-0.5'])
        self.result_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.setItemDelegateForColumn(0, ReadOnlyDelegate(self))  # Set the first column as read-only
        self.cell_layout = QtWidgets.QGridLayout()
        self.cell_buttons = []
        for i in range(4):
            for j in range(4):
                button = QtWidgets.QPushButton(f'Cell {i*4 + j + 1}')
                button.clicked.connect(lambda checked=False, row=i, col=j: self.cell_button_clicked(row, col))
                self.cell_layout.addWidget(button, i, j)
                self.cell_buttons.append(button)
        self.cell_save_button = QtWidgets.QPushButton('S')
        self.cell_save_button.clicked.connect(self.save_cell_data)
        self.cell_save_button.setStyleSheet("background-color: rgba(72, 250, 215, 51); border-radius: 10px; border: none;")
        self.cell_layout.addWidget(self.cell_save_button, 4, 3)
        self.right_layout.addLayout(self.cell_layout)
        self.layout.addWidget(self.table)
        self.layout.addLayout(self.right_layout)
        self.setLayout(self.layout)

        # Set the values for the first column of the parameter table
        self.param_table.setItem(0, 0, QtWidgets.QTableWidgetItem('k'))
        self.param_table.setItem(1, 0, QtWidgets.QTableWidgetItem('b'))
        self.param_table.setItem(2, 0, QtWidgets.QTableWidgetItem('Уход'))
        self.param_table.setItem(3, 0, QtWidgets.QTableWidgetItem('RnS'))
        self.param_table.setItem(4, 0, QtWidgets.QTableWidgetItem('Ошибка'))

        # Set background color for RnS, Rn^-0.5 columns in the data table
        rn_s_column = 2
        rn_sqrt_column = 5
        rn_s_delegate = BackgroundDelegate(QtGui.QColor(240, 240, 240, 128))
        rn_sqrt_delegate = BackgroundDelegate(QtGui.QColor(240, 240, 240, 128))
        self.table.setItemDelegateForColumn(rn_s_column, rn_s_delegate)
        self.table.setItemDelegateForColumn(rn_sqrt_column, rn_sqrt_delegate)

        # Set background color for Value column in the result table
        value_column = 1
        value_delegate = BackgroundDelegate(QtGui.QColor(240, 240, 240, 128))
        self.result_table.setItemDelegateForColumn(value_column, value_delegate)

        # Set background color for Number column in the data table
        number_column = 0
        number_delegate = BackgroundDelegate(QtGui.QColor(192, 192, 192, 128))
        self.table.setItemDelegateForColumn(number_column, number_delegate)

        # Set background color for Parameter column in the result table
        parameter_column = 0
        parameter_delegate = BackgroundDelegate(QtGui.QColor(192, 192, 192, 128))
        self.param_table.setItemDelegateForColumn(parameter_column, parameter_delegate)

    def calculate_results(self):
        self.update_plot()
        # Update the RnS column in the table
        for row in range(self.table.rowCount()):
            item_diameter = self.table.item(row, 3)
            item_rn = self.table.item(row, 4)
            if item_diameter is not None and item_diameter.text() and item_rn is not None and item_rn.text():
                try:
                    diameter = float(item_diameter.text())
                    rn = float(item_rn.text())
                    zero_x = float(self.param_table.item(2, 1).text())  # Get the Уход value from the parameter table
                    rns_value = rn * 0.25 * math.pi * (diameter - zero_x) ** 2
                    self.table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(round(rns_value, 4))))
                except ValueError:
                    pass
        # Update the result table
        self.update_result_table()

    def update_plot(self):
        x = []
        y = []
        for i in range(self.table.rowCount()):
            item_x = self.table.item(i, 3)
            item_y = self.table.item(i, 5)
            if item_x is not None and item_y is not None and item_x.text() and item_y.text():
                try:
                    x.append(float(item_x.text()))
                    y.append(float(item_y.text()))
                except ValueError:
                    pass
        self.plot.clear()
        self.plot.plot(x, y, name="Data", symbolSize=6, symbolBrush="#088F8F")
        if len(x) > 1 and len(y) > 1:
            b, a = linear_fit(x, y)
            zero_x = -a / b
            if np.min(x) > zero_x:
                x.insert(0, zero_x)
            if np.max(x) < zero_x:
                x.append(zero_x)
            y_appr = np.vectorize(lambda x: linear(x, b, a))(x)
            pen2 = pg.mkPen(color="#FF0000", width=3)
            self.plot.plot(x, y_appr, name="Fit", pen=pen2, symbolSize=0, symbolBrush=pen2.color())
            self.param_table.setItem(0, 1, QtWidgets.QTableWidgetItem(str(round(b, 4))))
            self.param_table.setItem(1, 1, QtWidgets.QTableWidgetItem(str(round(a, 4))))
            self.param_table.setItem(2, 1, QtWidgets.QTableWidgetItem(str(round(zero_x, 4))))
            rns = math.pi * 0.25 / (b ** 2)
            self.param_table.setItem(3, 1, QtWidgets.QTableWidgetItem(str(round(rns, 4))))

            # Calculate the mean deviation from the mean of the approximation
            y_mean = np.mean(y)
            deviation = np.mean(np.abs(y - y_mean))
            self.param_table.setItem(4, 1, QtWidgets.QTableWidgetItem(str(round(deviation, 4))))

    def prepare_plot(self):
        self.plot.setBackground("w")
        self.plot.setTitle(self.plot_title, color="#413C58", size="10pt")
        styles = {"color": "#413C58", "font-size": "15px"}
        self.plot.setLabel("left", self.y_label, **styles)
        self.plot.setLabel("bottom", self.x_label, **styles)
        self.plot.addLegend()
        self.plot.showGrid(x=True, y=True)

    def save_data(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        file_name, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Data", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_name:
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Data"
            headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            ws1.append(headers)
            for row in range(self.table.rowCount()):
                data = [self.table.item(row, col).text() for col in range(self.table.columnCount())]
                ws1.append(data)
            ws2 = wb.create_sheet("Results")
            headers = [self.param_table.horizontalHeaderItem(i).text() for i in range(self.param_table.columnCount())]
            ws2.append(headers)
            for row in range(self.param_table.rowCount()):
                data = [self.param_table.item(row, col).text() for col in range(self.param_table.columnCount())]
                ws2.append(data)
            wb.save(filename=file_name)

    def clean_rn(self):
        for row in range(self.table.rowCount()):
            self.table.setItem(row, 2, QtWidgets.QTableWidgetItem(''))  # Clear RnS column
            self.table.setItem(row, 4, QtWidgets.QTableWidgetItem(''))  # Clear Resistance column
            self.table.setItem(row, 5, QtWidgets.QTableWidgetItem(''))  # Clear Rn^-0.5 column
        self.plot.clear()

    def clean_all(self):
        for row in range(self.table.rowCount()):
            self.table.setItem(row, 1, QtWidgets.QTableWidgetItem(''))  # Clear Name column
            self.table.setItem(row, 2, QtWidgets.QTableWidgetItem(''))  # Clear RnS column
            self.table.setItem(row, 3, QtWidgets.QTableWidgetItem(''))  # Clear Diameter column
            self.table.setItem(row, 4, QtWidgets.QTableWidgetItem(''))  # Clear Resistance column
            self.table.setItem(row, 5, QtWidgets.QTableWidgetItem(''))  # Clear Rn^-0.5 column
        self.plot.clear()

    def cell_button_clicked(self, row, col):
        series, ok = QtWidgets.QInputDialog.getText(self, 'Input Dialog', 'Enter series:')
        if ok:
            uhod = self.param_table.item(2, 1).text() if self.param_table.item(2, 1) is not None else ''
            rns = self.param_table.item(3, 1).text() if self.param_table.item(3, 1) is not None else ''
            button = self.cell_buttons[row * 4 + col]
            button.setText(f'Серия: {series}\nУход: {uhod}\nRnS: {rns}')

    def save_cell_data(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        file_name, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save Cell Data", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_name:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cell Data"
            headers = ['Cell', 'Series', 'Уход', 'RnS']
            ws.append(headers)
            for i, button in enumerate(self.cell_buttons):
                text = button.text()
                if text:
                    series = text.split('\n')[0].split(': ')[1]
                    uhod = text.split('\n')[1].split(': ')[1]
                    rns = text.split('\n')[2].split(': ')[1]
                    data = [f'Cell {i+1}', series, uhod, rns]
                    ws.append(data)
            wb.save(filename=file_name)

    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_Enter or event.key() == QtCore.Qt.Key_Return:
            row = self.table.currentRow()
            col = self.table.currentColumn()
            if row < self.table.rowCount() - 1:
                self.table.setCurrentCell(row + 1, col)
        elif event.key() == QtCore.Qt.Key_Delete:
            selected_items = self.table.selectedItems()
            if selected_items:
                for item in selected_items:
                    if item.column() not in [0, 2, 5]:  # Disable delete for columns 0, 2, and 5
                        self.table.setItem(item.row(), item.column(), QtWidgets.QTableWidgetItem(''))
                self.update_plot()
        else:
            super(Window, self).keyPressEvent(event)

    def update_result_table(self):
        # Update the result table with the calculated values
        self.result_table.setRowCount(self.table.rowCount())
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item is not None:
                    self.result_table.setItem(row, col, QtWidgets.QTableWidgetItem(item.text()))

app = QtWidgets.QApplication(sys.argv)
window = Window()
window.show()
sys.exit(app.exec_())
