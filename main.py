import logging
import re
import sys

import numpy as np
import openpyxl
import pyqtgraph as pg
from openpyxl.styles import Alignment, Border, Font, Side
from PySide6 import QtWidgets
from PySide6.QtGui import QIcon

from constants import BLACK, PLOT_COLORS, RNS_ERROR_COLOR, WHITE, DataTableColumns, ParamTableColumns
from errors import ListsNotSameLength
from store import InitialDataItem, InitialDataItemList, Store
from utils import (
    calculate_drift,
    calculate_drift_per_sample,
    calculate_rn_sqrt,
    calculate_rns,
    calculate_rns_error_diff,
    calculate_rns_error_per_sample,
    calculate_rns_per_sample,
    calculate_square,
    drop_nans,
    linear,
    linear_fit,
)
from widgets.cell import CellWidget
from widgets.tables.data_table import DataTable
from widgets.tables.item import TableWidgetItem
from widgets.tables.param_table import ParamTable

logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler("debug.log")],
)


class Window(QtWidgets.QWidget):
    def __init__(self) -> None:
        super(Window, self).__init__()
        self.setGeometry(100, 100, 1200, 900)

        self.setWindowIcon(QIcon("./assets/rns-logo-sm.png"))
        # Таблица с исходными данными
        self.data_table_label = QtWidgets.QLabel("Таблица с данными", self)
        self.data_table = DataTable(rows=50)

        # График
        self.plot = pg.PlotWidget()
        self.prepare_plot()

        # Main Layout
        self.layout = QtWidgets.QHBoxLayout()

        # Left Layout (таблица с данными)
        self.left_layout = QtWidgets.QVBoxLayout()

        # Right Layout (таблица с параметрами, график, экшнс)
        self.right_layout = QtWidgets.QVBoxLayout()

        # Right layout параметры расчета
        self.actions_params_layout = QtWidgets.QHBoxLayout()

        # Таблица с параметрами
        self.param_table_label = QtWidgets.QLabel("Таблица с расчетом", self)
        self.param_table = ParamTable()

        # Экшнс кнопки
        self.actions_group = QtWidgets.QGroupBox("Действия")
        self.actions_layout = QtWidgets.QHBoxLayout()

        self.result_button = QtWidgets.QPushButton("Расcчитать")
        self.result_button.setToolTip("Произвести расчеты")
        self.result_button.clicked.connect(self.calculate_results)

        self.clean_rn_button = QtWidgets.QPushButton("Очистить Rn")
        self.clean_rn_button.setToolTip("Очистить стоблец Rn и таблицу с расчетом")
        self.clean_rn_button.clicked.connect(self.clean_rn)

        self.clean_all_button = QtWidgets.QPushButton("Очистить таблицы")
        self.clean_all_button.setToolTip("Очистить график и таблицы с даными и расчетом")
        self.clean_all_button.clicked.connect(self.clean_all)

        self.actions_layout.addWidget(self.result_button)
        self.actions_layout.addWidget(self.clean_rn_button)
        self.actions_layout.addWidget(self.clean_all_button)
        self.actions_group.setLayout(self.actions_layout)

        # Параметры для расчета

        self.rn_consistent_group = QtWidgets.QGroupBox("Последовательное Rn (Ом)")
        self.rn_consistent_layout = QtWidgets.QHBoxLayout()
        self.rn_consistent = QtWidgets.QDoubleSpinBox(self)
        self.rn_consistent.setRange(0, 100)
        self.rn_consistent.setDecimals(2)
        self.rn_consistent.setValue(0)
        self.rn_consistent_layout.addWidget(self.rn_consistent)
        self.rn_consistent_group.setLayout(self.rn_consistent_layout)

        self.allowed_error_group = QtWidgets.QGroupBox("Разрешенная ошибка")
        self.allowed_error_layout = QtWidgets.QHBoxLayout()
        self.allowed_error = QtWidgets.QDoubleSpinBox(self)
        self.allowed_error.setRange(0, 1)
        self.allowed_error.setDecimals(2)
        self.allowed_error.setValue(0)
        self.allowed_error_layout.addWidget(self.allowed_error)
        self.allowed_error_group.setLayout(self.allowed_error_layout)

        # Грид с ячейками
        self.cell_group = QtWidgets.QGroupBox("Запись")
        self.cell_v_layout = QtWidgets.QVBoxLayout()
        self.cell_h_layout = QtWidgets.QHBoxLayout()
        self.cell_grid_layout = QtWidgets.QGridLayout()
        self.cell_widgets = []
        for i in range(4):
            for j in range(4):
                index = i * 4 + j + 1
                cell_widget = CellWidget(self, index, self.param_table)
                self.cell_grid_layout.addWidget(cell_widget, i, j)
                self.cell_widgets.append(cell_widget)

        self.mean_drift = QtWidgets.QLabel("Средний уход: --", self)
        self.mean_rns = QtWidgets.QLabel("Средний RnS: --", self)

        self.btn_load_cell_data = QtWidgets.QPushButton("Загрузить из файла")
        self.btn_load_cell_data.setToolTip("Загрузить данные из xlsx файла")
        self.btn_load_cell_data.clicked.connect(self.load_cell_data)

        self.btn_clear_cell_data = QtWidgets.QPushButton("Очистить ячейки")
        self.btn_clear_cell_data.setToolTip("Очистить ячейки с записанными данными")
        self.btn_clear_cell_data.clicked.connect(self.clear_cell_data)

        self.cell_save_button = QtWidgets.QPushButton("Сохранить")
        self.cell_save_button.setToolTip("Сохранить записанные данные с RnS")
        self.cell_save_button.clicked.connect(self.save_cell_data)

        self.cell_h_layout.addWidget(self.mean_drift)
        self.cell_h_layout.addWidget(self.mean_rns)
        self.cell_h_layout.addWidget(self.btn_clear_cell_data)
        self.cell_h_layout.addWidget(self.btn_load_cell_data)
        self.cell_h_layout.addWidget(self.cell_save_button)

        self.cell_v_layout.addLayout(self.cell_grid_layout)
        self.cell_v_layout.addLayout(self.cell_h_layout)
        self.cell_group.setLayout(self.cell_v_layout)

        # Добавляем виджеты в левый лейаут
        self.left_layout.addWidget(self.data_table_label)
        self.left_layout.addWidget(self.data_table)

        # Добавляем все виджеты в правый лейаут
        self.right_layout.addWidget(self.param_table_label)
        self.right_layout.addWidget(self.param_table)
        self.right_layout.addWidget(self.plot)
        self.actions_params_layout.addWidget(self.actions_group)
        self.actions_params_layout.addWidget(self.rn_consistent_group)
        self.actions_params_layout.addWidget(self.allowed_error_group)
        self.right_layout.addLayout(self.actions_params_layout)
        self.right_layout.addWidget(self.cell_group)

        # Добавляет виджеты в основной лейаут
        self.layout.addLayout(self.left_layout)
        self.layout.addLayout(self.right_layout)
        self.setLayout(self.layout)

    def calculate_means(self):
        rns_list = []
        drift_list = []
        for cell in self.cell_widgets:
            _, drift, rns = self.parse_cell(cell)
            if drift:
                drift_list.append(drift)
            if rns:
                rns_list.append(rns)

        self.mean_drift.setText(f"Средний уход: {round(np.mean(drift_list), 3)}")
        self.mean_rns.setText(f"Средний RnS: {round(np.mean(rns_list), 1)}")

    def clear_means(self):
        self.mean_drift.setText("Средний уход: --")
        self.mean_rns.setText("Средний RnS: --")

    def calculate_rn05(self):
        """Расчет Rn^-0.5 для каждого образца"""
        for row in range(self.data_table.rowCount()):
            resistance = self.data_table.get_column_value(row, DataTableColumns.RESISTANCE)
            diameter = self.data_table.get_column_value(row, DataTableColumns.DIAMETER)
            if not resistance or not diameter:
                continue
            rn_sqrt = calculate_rn_sqrt(resistance=resistance, rn_consistent=self.rn_consistent.value())
            self.data_table.setItem(row, DataTableColumns.RN_SQRT.index, TableWidgetItem(str(rn_sqrt)))
        return True

    def calculate_main_params(self):
        """Расчет Наклона, Пересечения, RnS, Ухода в целом"""
        diameter_list = self.data_table.get_column_values(DataTableColumns.DIAMETER)
        rn_sqrt_list = self.data_table.get_column_values(DataTableColumns.RN_SQRT)
        try:
            diameter_list, rn_sqrt_list = drop_nans(diameter_list, rn_sqrt_list)
        except (ListsNotSameLength, ValueError):
            QtWidgets.QMessageBox.warning(
                self,
                "Не корректные данные!",
                "В таблице данных не корректные значения для 'Диаметр ACAD' и 'Rn^-0.5'",
            )
            return False
        slope, intercept = linear_fit(diameter_list, rn_sqrt_list)

        self.param_table.setItem(
            0,
            ParamTableColumns.SLOPE.index,
            TableWidgetItem(str(slope)),
        )
        self.param_table.setItem(
            0,
            ParamTableColumns.INTERCEPT.index,
            TableWidgetItem(str(intercept)),
        )

        drift = calculate_drift(slope=slope, intercept=intercept)

        self.param_table.setItem(
            0,
            ParamTableColumns.DRIFT.index,
            TableWidgetItem(str(drift)),
        )

        rns = calculate_rns(slope)
        self.param_table.setItem(
            0,
            ParamTableColumns.RNS.index,
            TableWidgetItem(str(rns)),
        )
        return True

    def calculate_error_params(self):
        """Расчет ошибок RnS и Ухода"""
        rns = self.param_table.get_column_value(0, ParamTableColumns.RNS)
        rns_list = np.array([v for v in self.data_table.get_column_values(DataTableColumns.RNS) if v], dtype=float)

        try:
            rns_error = np.sqrt(np.sum((rns_list - rns) ** 2) / len(rns_list))
        except (ZeroDivisionError,):
            QtWidgets.QMessageBox.warning(
                self,
                "Ошибка в рассчете RNS_ERROR",
                "Ошибка деления на ноль при расчете RNS_ERROR!",
            )
            return False

        self.param_table.setItem(
            0,
            ParamTableColumns.RNS_ERROR.index,
            TableWidgetItem(str(rns_error)),
        )

        for row in range(self.data_table.rowCount()):
            self.data_table.color_row(row=row, background_color=WHITE, text_color=BLACK)
            rns_value = self.data_table.get_column_value(row, DataTableColumns.RNS)
            if not rns_value:
                continue
            value = calculate_rns_error_per_sample(rns_i=rns_value, rns=rns)
            self.data_table.setItem(row, DataTableColumns.RNS_ERROR.index, TableWidgetItem(str(value)))
            error_diff = calculate_rns_error_diff(
                rns_error_per_sample=value, rns_error=rns_error, allowed_error=self.allowed_error.value()
            )
            if error_diff > 0:
                self.data_table.color_row(row=row, background_color=RNS_ERROR_COLOR, text_color=WHITE)
            else:
                self.data_table.color_row(row=row, background_color=WHITE, text_color=BLACK)

        drift = self.param_table.get_column_value(0, ParamTableColumns.DRIFT)
        drift_list = np.array([v for v in self.data_table.get_column_values(DataTableColumns.DRIFT) if v], dtype=float)
        drift_error = np.sqrt(np.sum((drift_list - drift) ** 2) / len(drift_list))
        self.param_table.setItem(
            0,
            ParamTableColumns.DRIFT_ERROR.index,
            TableWidgetItem(str(drift_error)),
        )

        self.param_table.setItem(
            0, ParamTableColumns.RN_CONSISTENT.index, TableWidgetItem(str(self.rn_consistent.value()))
        )
        self.param_table.setItem(
            0, ParamTableColumns.ALLOWED_ERROR.index, TableWidgetItem(str(self.allowed_error.value()))
        )
        return True

    def calculate_rns_drift_square_per_sample(self):
        """Расчет RnS, Ухода и Площади для каждого образца по отдельности"""
        drift = self.param_table.get_column_value(0, ParamTableColumns.DRIFT)
        if not drift:
            QtWidgets.QMessageBox.warning(
                self,
                "Не корректные данные!",
                "Не вычислен уход!",
            )
            return False

        rns_mean = self.param_table.get_column_value(0, ParamTableColumns.RNS)
        if not rns_mean:
            QtWidgets.QMessageBox.warning(
                self,
                "Не корректные данные!",
                "Не вычислен RnS!",
            )
            return False

        for row in range(self.data_table.rowCount()):
            diameter = self.data_table.get_column_value(row, DataTableColumns.DIAMETER)
            resistance = self.data_table.get_column_value(row, DataTableColumns.RESISTANCE)
            if not resistance or not diameter:
                continue

            square_value = calculate_square(diameter=diameter, drift=drift)  # подставлям общий уход
            self.data_table.setItem(row, DataTableColumns.SQUARE.index, TableWidgetItem(str(square_value)))

            rns_value = calculate_rns_per_sample(
                resistance=resistance, diameter=diameter, drift=drift, rn_persistent=self.rn_consistent.value()
            )
            self.data_table.setItem(
                row,
                DataTableColumns.RNS.index,
                TableWidgetItem(str(rns_value)),
            )

            drift_value = calculate_drift_per_sample(
                diameter=diameter, resistance=resistance, rns=rns_mean, rn_persistent=self.rn_consistent.value()
            )
            self.data_table.setItem(row, DataTableColumns.DRIFT.index, TableWidgetItem(str(drift_value)))
        return True

    def calculate_results(self):
        self.data_table.clear_calculations()
        if not self.calculate_rn05():
            return
        if not self.calculate_main_params():
            return
        if not self.calculate_rns_drift_square_per_sample():
            return
        if not self.calculate_error_params():
            return
        self.plot_current_data()

    def plot_current_data(self):
        diameter_list = self.data_table.get_column_values(DataTableColumns.DIAMETER)
        rn_sqrt_list = self.data_table.get_column_values(DataTableColumns.RN_SQRT)
        try:
            diameter_list, rn_sqrt_list = drop_nans(diameter_list, rn_sqrt_list)
            diameter_list, rn_sqrt_list = np.array(
                sorted(np.array([diameter_list, rn_sqrt_list]).T, key=lambda x: x[0]), dtype=float
            ).T
            diameter_list = diameter_list.tolist()
            rn_sqrt_list = rn_sqrt_list.tolist()
        except (ListsNotSameLength, ValueError):
            return False
        drift = self.param_table.get_column_value(0, ParamTableColumns.DRIFT)
        slope = self.param_table.get_column_value(0, ParamTableColumns.SLOPE)
        intercept = self.param_table.get_column_value(0, ParamTableColumns.INTERCEPT)

        plotItem = self.plot.getPlotItem()
        items_data = [item for item in plotItem.items if item.name() == "Data"]

        items_fit = [item for item in plotItem.items if item.name() == "Fit"]

        if items_data:
            items_data[0].setData(diameter_list, rn_sqrt_list)
        else:
            self.plot.plot(diameter_list, rn_sqrt_list, name="Data", symbolSize=6, symbolBrush="#000000")

        if np.min(diameter_list) > drift:
            diameter_list.insert(0, drift)
        if np.max(diameter_list) < drift:
            diameter_list.append(drift)
        y_appr = np.vectorize(lambda x: linear(x, slope, intercept))(diameter_list)

        if items_fit:
            items_fit[0].setData(diameter_list, y_appr)
            return None

        else:
            pen2 = pg.mkPen(color="#000000", width=3)
            self.plot.plot(
                diameter_list,
                y_appr,
                name="Fit",
                pen=pen2,
                symbolSize=0,
                symbolBrush=pen2.color(),
            )
            return None

    def addCellData(self, cell: int, name: str):
        Store.update_or_create_item(
            cell=cell,
            name=name,
            diameter_list=self.data_table.get_column_values(DataTableColumns.DIAMETER),
            rn_sqrt_list=self.data_table.get_column_values(DataTableColumns.RN_SQRT),
            slope=self.param_table.get_column_value(0, ParamTableColumns.SLOPE),
            intercept=self.param_table.get_column_value(0, ParamTableColumns.INTERCEPT),
            drift=self.param_table.get_column_value(0, ParamTableColumns.DRIFT),
            rns=self.param_table.get_column_value(0, ParamTableColumns.RNS),
            drift_error=self.param_table.get_column_value(0, ParamTableColumns.DRIFT_ERROR),
            rns_error=self.param_table.get_column_value(0, ParamTableColumns.RNS_ERROR),
            initial_data=self.data_table.dump_data(),
            rn_consistent=self.rn_consistent.value(),
            allowed_error=self.allowed_error.value(),
        )
        self.set_active_cell(cell)

    def plot_data(self, cell: int):
        color = PLOT_COLORS[cell - 1]
        item = Store.data.get(cell=cell)
        if not item:
            return
        diameter, rn_sqrt = drop_nans(item.diameter_list, item.rn_sqrt_list)
        diameter_list = diameter.tolist()
        if np.min(diameter_list) > item.drift:
            diameter_list.insert(0, item.drift)
        if np.max(diameter_list) < item.drift:
            diameter_list.append(item.drift)
        y_appr = np.vectorize(lambda x: linear(x, item.slope, item.intercept))(diameter_list)
        pen2 = pg.mkPen(color=color, width=2)
        self.plot.plot(
            diameter_list,
            y_appr,
            name=f"{item.name}",
            pen=pen2,
            symbolSize=0,
            symbolBrush=pen2.color(),
        )

    def remove_plot(self, cell: int):
        cell_data = Store.data.get(cell=cell)
        plotItem = self.plot.getPlotItem()
        items_to_remove = [item for item in plotItem.items if item.name() == cell_data.name]
        for item in items_to_remove:
            plotItem.removeItem(item)

    def prepare_plot(self):
        y_label = "Rn^-0.5"
        x_label = "Диаметр ACAD (μm)"
        self.plot.setBackground("w")
        styles = {"color": "#413C58", "font-size": "15px"}
        self.plot.setLabel("left", y_label, **styles)
        self.plot.setLabel("bottom", x_label, **styles)
        self.plot.addLegend()
        self.plot.showGrid(x=True, y=True)

    def clean_rn(self):
        self.data_table.clear_rn()
        self.param_table.clear_all()

    def clean_all(self):
        self.data_table.clear_all()
        self.param_table.clear_all()
        self.plot.clear()
        self.set_active_cell(0)

    @staticmethod
    def parse_cell(cell):
        drift = float(cell.drift.text().split("Уход: ")[1]) if cell.drift.text() else ""
        rns = float(cell.rns.text().split("RnS: ")[1]) if cell.rns.text() else ""
        return [cell.name.text(), drift, rns]

    def save_cell_data(self):
        options = QtWidgets.QFileDialog.Options()
        options |= QtWidgets.QFileDialog.DontUseNativeDialog
        file_name, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Cell Data",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options,
        )
        if not file_name:
            return

        wb = openpyxl.Workbook()
        ws_cells = wb.active
        ws_cells.title = "Cells data"

        init_data = [[cell.name.text(), cell.drift.text(), cell.rns.text()] for cell in self.cell_widgets]
        output = []

        # Разделение исходного массива на блоки по 4 строки
        blocks = [init_data[i : i + 4] for i in range(0, len(init_data), 4)]

        # Обработка каждого блока
        for block in blocks:
            # Перестановка элементов в блоке
            block_transposed = list(map(list, zip(*block)))
            # Добавление переставленного блока в выходной массив
            output.extend(block_transposed)
        for row_ind, row in enumerate(output, 1):
            for col_ind, coll in enumerate(row, 1):
                cell = ws_cells.cell(row=row_ind, column=col_ind, value=coll)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if (row_ind - 1) % 3 == 0:  # для клеток с названием серии
                    cell.border = Border(
                        right=Side(style="thick"),
                        left=Side(style="thick"),
                        top=Side("thick"),
                        bottom=Side(style="thick"),
                    )
                    cell.font = Font(bold=True)
                elif (row_ind - 3) % 3 == 0:  # для клеток с названием rns
                    cell.border = Border(
                        right=Side(style="thick"), left=Side(style="thick"), bottom=Side(style="thick")
                    )
                else:  # для остальных клеток
                    cell.border = Border(right=Side(style="thick"), left=Side(style="thick"))

        # Устанавливаем ширину всех столбцов
        for col in ws_cells.columns:
            column = col[0].column_letter  # Получаем букву столбца
            ws_cells.column_dimensions[column].width = 12

        # Устанавливаем высоту для всех строк
        for row in ws_cells.rows:
            ws_cells.row_dimensions[row[0].row].height = 21

        # Сохраняем все данные
        data_headers = DataTableColumns.get_all_slugs()
        results_headers = ParamTableColumns.get_all_names()
        for cell_data in Store.data:
            ws_data = wb.create_sheet(f"Data №{cell_data.cell} {cell_data.name}")
            ws_data.append(data_headers)
            for dat in cell_data.initial_data:
                ws_data.cell(row=dat["row"] + 2, column=dat["col"] + 1, value=dat["value"])

            ws_results = wb.create_sheet(f"Results №{cell_data.cell} {cell_data.name}")
            ws_results.append(results_headers)
            for i, param in enumerate(ParamTableColumns):
                ws_results.cell(row=2, column=i + 1, value=getattr(cell_data, param.slug, ""))

        # Save the Excel file
        if not file_name.endswith(".xlsx"):
            file_name += ".xlsx"
        wb.save(filename=file_name)

    def reload_tables_from_cell_data(self, cell: int):
        cell_data = Store.data.get(cell=cell)
        if not cell_data:
            return
        self.data_table.load_data(data=cell_data.initial_data)
        self.param_table.load_data(data=cell_data)
        self.plot_current_data()
        self.set_active_cell(cell)
        self.rn_consistent.setValue(cell_data.rn_consistent)
        self.allowed_error.setValue(cell_data.allowed_error)

    def set_active_cell(self, cell: int):
        for cw in self.cell_widgets:
            if cw.index == cell:
                cw.set_active(True)
                continue
            cw.set_active(False)

    def clear_cell_data(self):
        reply = QtWidgets.QMessageBox.question(
            self,
            "Очистить записанные данные ячеек",
            "Записанные данные ячеек будут удалены, продолжить?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )

        if reply == QtWidgets.QMessageBox.Yes:
            for cell in self.cell_widgets:
                cell.clear()
            Store.clear()
            self.clear_means()
            self.set_active_cell(0)

    def load_cell_data(self):
        reply = QtWidgets.QMessageBox.question(
            self,
            "Загрузка данных их файла",
            "Все текущие данные будут очищены, продолжить?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )

        if reply != QtWidgets.QMessageBox.Yes:
            return

        for cell in self.cell_widgets:
            cell.clear()
        Store.clear()
        self.clear_means()
        self.clean_all()
        self.set_active_cell(0)

        options = QtWidgets.QFileDialog.Options()
        fileName, _ = QtWidgets.QFileDialog.getOpenFileName(
            None, "Выберите файл XLSX", "", "Excel Files (*.xlsx);;All Files (*)", options=options
        )

        if not fileName:
            return

        is_some_errors = False
        some_errors_text = ""
        try:
            wb = openpyxl.load_workbook(fileName)
            data_sheet_names = [sh for sh in wb.sheetnames if sh.startswith("Data №")]
            if not len(data_sheet_names):
                QtWidgets.QMessageBox.critical(
                    self, "Ошибка чтения", "Не найдены данные с нумерацией для записанных ячеек"
                )
                return
            sheet_names = "\n".join(wb.sheetnames)
            for sheet_name in data_sheet_names:
                try:
                    i = int(re.findall(r"Data №(\d+) .*", sheet_name)[0])
                    data_name = re.findall(f"(Data №{i} .*)", sheet_name)[0]
                    result_name = re.findall(f"(Results №{i} .*)", sheet_names)[0]
                    cell_name = re.findall(f"Data №{i} (.*)", sheet_name)[0]
                except (IndexError, ValueError):
                    is_some_errors = True
                    continue

                ws_data = wb[data_name]
                ws_result = wb[result_name]

                initial_data = InitialDataItemList()
                column_names = [ws_data[1][col].value for col in range(0, ws_data.max_column)]
                for data_column in DataTableColumns:
                    col = None
                    try:
                        col = column_names.index(data_column.slug)
                    except (ValueError,):
                        is_some_errors = True
                        some_errors_text += f"\nКолонка '{data_column.slug}' не найдена в таблице '{data_name}';"

                    for row in range(2, ws_data.max_row + 1):
                        try:
                            value = ws_data[row][col].value if col is not None and ws_data[row][col].value else ""
                            initial_data.append(InitialDataItem(row=row - 2, col=data_column.index, value=value))
                        except IndexError:
                            is_some_errors = True
                            continue

                diameter_list = [
                    float(v.value) if v.value else None
                    for v in initial_data.filter(col=DataTableColumns.DIAMETER.index)
                ]
                rn_sqrt_list = [
                    float(v.value) if v.value else None for v in initial_data.filter(col=DataTableColumns.RN_SQRT.index)
                ]

                result_column_names = [ws_result[1][col].value for col in range(0, ws_result.max_column)]
                result_kwargs = {}
                for result_column in ParamTableColumns:
                    try:
                        col = result_column_names.index(result_column.name)
                    except (ValueError,):
                        is_some_errors = True
                        some_errors_text += f"\nКолонка '{result_column.name}' не найдена в таблице '{result_name}';"
                        continue
                    row = 2
                    try:
                        value = ws_result[row][col].value if col is not None and ws_result[row][col].value else 0
                        result_kwargs[result_column.slug] = value
                    except IndexError:
                        is_some_errors = True
                        continue

                cell_item = Store.update_or_create_item(
                    cell=i,
                    name=cell_name,
                    diameter_list=diameter_list,
                    rn_sqrt_list=rn_sqrt_list,
                    initial_data=initial_data,
                    **result_kwargs,
                )

                cell_widget = self.cell_widgets[cell_item.cell - 1]
                cell_widget.name.setText(cell_item.name)
                cell_widget.drift.setText(f"Уход: {round(cell_item.drift, 3)}")
                cell_widget.rns.setText(f"RnS: {round(cell_item.rns, 1)}")
                cell_widget.updateUI()
                self.calculate_means()

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Ошибка чтения", f"Возникли ошибки чтения файла: {str(e)}")
            logger.exception(f"{e}", exc_info=True)
            return

        if is_some_errors:
            QtWidgets.QMessageBox.warning(
                self,
                "Файл прочитался с ошибками",
                some_errors_text,
            )
            return

        QtWidgets.QMessageBox.information(self, "Файл успешно прочитан", "Все данные успешно загружены")


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = Window()
    window.show()
    sys.exit(app.exec_())
